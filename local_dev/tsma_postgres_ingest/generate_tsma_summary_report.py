#!/usr/bin/env python3
"""
Generate a simple TSMA Excel summary report from Postgres.

The workbook contains one sheet with monthly columns from Jan 2024 to Dec 2026.
Rows:
- Cellular
- Data
- Voice
- Out of Scope
- Total

Then one block per entity/category with the same rows.
"""

from __future__ import annotations

import argparse
import os
from decimal import Decimal
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from tsma_summary_report_utils import (
    DATA_TOWERS,
    MONTH_CODES,
    MONTH_HEADERS,
    OUT_OF_SCOPE_TOWERS,
    VOICE_TOWERS,
    blank_amounts,
    entity_names,
    load_cellular,
    load_ivr_totals,
    load_lite_cellular,
    load_lite_wireline,
    load_wireline,
    merge_maps,
    sum_maps,
)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a TSMA monthly summary workbook.")
    parser.add_argument("--dsn", default=os.environ.get("DATABASE_URL"), help="Postgres DSN. Defaults to DATABASE_URL.")
    parser.add_argument(
        "--output",
        default="local_dev/tsma_postgres_ingest/tsma_summary_2024_2026.xlsx",
        help="Output .xlsx path.",
    )
    return parser.parse_args()


def write_amount_row(ws, row_num: int, label: str, amounts: list[Decimal], bold: bool = False) -> None:
    ws.cell(row=row_num, column=2, value=label)
    ws.cell(row=row_num, column=2).font = Font(bold=bold)
    ws.cell(row=row_num, column=2).border = THIN_BORDER
    for col_num, amount in enumerate(amounts, start=3):
        cell = ws.cell(row=row_num, column=col_num, value=float(amount))
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=bold)
        cell.border = THIN_BORDER


def write_total_row(ws, row_num: int, source_rows: list[int]) -> None:
    ws.cell(row=row_num, column=2, value="Total")
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    ws.cell(row=row_num, column=2).border = THIN_BORDER
    for col_num in range(3, 3 + len(MONTH_CODES)):
        refs = [f"{get_column_letter(col_num)}{row}" for row in source_rows]
        cell = ws.cell(row=row_num, column=col_num, value=f"=SUM({','.join(refs)})")
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER


def style_sheet(ws) -> None:
    ws.title = "TSMA"
    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    for col_num in range(3, 3 + len(MONTH_CODES)):
        ws.column_dimensions[get_column_letter(col_num)].width = 12


def build_workbook(
    cellular_map: dict[str, list[Decimal]],
    data_map: dict[str, list[Decimal]],
    voice_map: dict[str, list[Decimal]],
    out_of_scope_map: dict[str, list[Decimal]],
    ivr_totals: list[Decimal],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    style_sheet(ws)

    ws["B1"] = "TSMA"
    ws["B1"].font = Font(bold=True)
    ws["B1"].border = THIN_BORDER
    for col_num, label in enumerate(MONTH_HEADERS, start=3):
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    row_num = 2
    write_amount_row(ws, row_num, "Cellular", sum_maps(cellular_map))
    write_amount_row(ws, row_num + 1, "Data", sum_maps(data_map))
    write_amount_row(ws, row_num + 2, "Voice", sum_maps(voice_map))
    write_amount_row(ws, row_num + 3, "Out of Scope", sum_maps(out_of_scope_map))
    write_total_row(ws, row_num + 4, [row_num, row_num + 1, row_num + 2, row_num + 3])

    row_num = 8
    for entity in entity_names(cellular_map, data_map, voice_map, out_of_scope_map):
        ws.cell(row=row_num, column=1, value=entity)
        ws.cell(row=row_num, column=1).font = Font(bold=True)

        write_amount_row(ws, row_num + 1, "Cellular", cellular_map.get(entity, blank_amounts()))
        write_amount_row(ws, row_num + 2, "Data", data_map.get(entity, blank_amounts()))
        write_amount_row(ws, row_num + 3, "Voice", voice_map.get(entity, blank_amounts()))

        total_rows = [row_num + 1, row_num + 2, row_num + 3]
        next_row = row_num + 4
        if entity == "GoBC":
            write_amount_row(ws, next_row, "Voice - IVR", ivr_totals)
            total_rows.append(next_row)
            next_row += 1

        write_amount_row(ws, next_row, "Out of Scope", out_of_scope_map.get(entity, blank_amounts()))
        total_rows.append(next_row)
        write_total_row(ws, next_row + 1, total_rows)
        row_num = next_row + 3

    return wb


def main() -> int:
    args = parse_args()
    if not args.dsn:
        raise SystemExit("Missing Postgres DSN. Pass --dsn or set DATABASE_URL.")

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)

    with psycopg.connect(args.dsn) as conn:
        # cellular_map = load_cellular(conn)
        # data_map = load_wireline(conn, DATA_TOWERS)
        # voice_map = load_wireline(conn, VOICE_TOWERS)
        # out_of_scope_map = load_wireline(conn, OUT_OF_SCOPE_TOWERS)
        cellular_map = merge_maps(load_cellular(conn), load_lite_cellular(conn))
        data_map = merge_maps(load_wireline(conn, DATA_TOWERS), load_lite_wireline(conn, DATA_TOWERS))
        voice_map = merge_maps(load_wireline(conn, VOICE_TOWERS), load_lite_wireline(conn, VOICE_TOWERS))
        out_of_scope_map = merge_maps(
            load_wireline(conn, OUT_OF_SCOPE_TOWERS),
            load_lite_wireline(conn, OUT_OF_SCOPE_TOWERS),
        )
        ivr_totals = load_ivr_totals(conn)

    workbook = build_workbook(cellular_map, data_map, voice_map, out_of_scope_map, ivr_totals)
    workbook.save(output)
    print(f"Wrote {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
