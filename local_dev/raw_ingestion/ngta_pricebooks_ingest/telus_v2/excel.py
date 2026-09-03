"""Parse Telus pricebook v2 Excel workbooks into per-table row dicts.

Uses openpyxl directly (not pandas) because a couple of sheets need merged
-cell awareness to fill down a category column correctly.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Callable, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from telus_v2.catalogues import (
    BookSpec,
    MultiBlockSheetSpec,
    SheetSpec,
    SimpleSheetSpec,
    SplitByValueSheetSpec,
    resolve_sheet,
)

HEADER_OVERRIDES: dict[str, str] = {
    "service_id": "service_id",
    "serviceid": "service_id",
    "service_id_for_ordering": "service_id",
    "service_id_billing_id": "service_id",
    "rate_plan": "rate_plan",
    "monthly_fee": "monthly_fee",
    "cpm_rate": "cpm_rate",
    "calling_to": "calling_to",
    "usage_rate": "usage_rate",
}


def canonical_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).replace(" ", " ").strip()
    if not s or s.lower().startswith("unnamed"):
        return ""
    s = re.sub(r"\s+", " ", s.lower())
    s = s.replace("(", " ").replace(")", " ").replace("*", " ")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if s in HEADER_OVERRIDES:
        return HEADER_OVERRIDES[s]
    # Long descriptive headers carry extra subtitle/qualifier text (e.g.
    # "Service SLA\nService Levels are described in...") that breaks an
    # exact match against the DB column name — fall back to substring
    # matching, same technique the old telus/excel.py used.
    if "landline" in s and "termination" in s and "cpm" in s:
        return "landline_termination_cpm_rate"
    if "mobile" in s and "termination" in s and "cpm" in s:
        return "mobile_termination_cpm_rate"
    if "sla" in s:
        return "service_sla"
    if "technical" in s and "support" in s:
        return "technical_services_support"
    if "technical" in s and "standard" in s:
        return "technical_service_standards"
    if "ordering" in s and "lead" in s:
        return "ordering_lead_times_objectives"
    if "delivery" in s and "lead" in s:
        return "delivery_lead_times_objectives_service_interval"
    return s


def as_text(value: Any, number_format: Optional[str] = None) -> Optional[str]:
    """Stringify a cell value. Numeric cells with a currency or percentage
    number format (e.g. Excel's "$5.00" or "25%" display) are rendered that
    way rather than as the bare underlying number (5, 0.25), since Telus
    enters real prices/rates as formatted numeric cells, not literal
    "$"/"%"-suffixed text like the old catalogues did."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool) and number_format:
        if "$" in number_format:
            return f"${value:,.2f}"
        if "%" in number_format:
            return f"{value * 100:g}%"
    s = str(value).strip()
    return s if s else None


def _row_values(ws: Worksheet, row: int, max_col: int) -> list[Any]:
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def _row_formats(ws: Worksheet, row: int, max_col: int) -> list[Optional[str]]:
    return [ws.cell(row=row, column=c).number_format for c in range(1, max_col + 1)]


def _cell_format(ws: Worksheet, row: int, col: int) -> Optional[str]:
    return ws.cell(row=row, column=col).number_format


def _is_blank_row(values: list[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _fill_down_merges(ws: Worksheet, values_by_row: dict[int, list[Any]], columns: list[int]) -> None:
    """Copy a merged range's top-left value into every row the range spans,
    for the given 1-indexed column numbers only."""
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col != merged_range.max_col:
            continue  # multi-column merges are section/footnote banners, not data to fill
        if merged_range.min_col not in columns:
            continue
        top_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            row_vals = values_by_row.get(r)
            if row_vals is not None:
                row_vals[merged_range.min_col - 1] = top_value


def _extract_header_rows(
    ws: Worksheet, header_row: int, columns: tuple[str, ...]
) -> list[dict[str, Any]]:
    """Shared header-row-plus-data-rows extraction used by both
    parse_simple_sheet (one table) and parse_split_by_value_sheet (several
    tables from one sheet)."""
    max_col = ws.max_column
    max_row = ws.max_row

    header_values = _row_values(ws, header_row, max_col)
    colmap: dict[int, str] = {}
    for idx, raw in enumerate(header_values):
        canon = canonical_header(raw)
        if canon:
            colmap[idx] = canon

    values_by_row: dict[int, list[Any]] = {
        r: _row_values(ws, r, max_col) for r in range(header_row + 1, max_row + 1)
    }
    formats_by_row: dict[int, list[Optional[str]]] = {
        r: _row_formats(ws, r, max_col) for r in range(header_row + 1, max_row + 1)
    }

    fill_cols = [c + 1 for c in colmap if colmap[c] == "category"]
    if fill_cols:
        _fill_down_merges(ws, values_by_row, fill_cols)

    db_cols = frozenset(columns)
    rows_out: list[dict[str, Any]] = []
    for r in sorted(values_by_row):
        raw_values = values_by_row[r]
        raw_formats = formats_by_row[r]
        if _is_blank_row(raw_values):
            continue
        vals: dict[str, Any] = {"pricebook_ingestion_run_id": None, "excel_row_number": r, "extras": None}
        extras: dict[str, Any] = {}
        for idx, canon in colmap.items():
            fmt = raw_formats[idx] if idx < len(raw_formats) else None
            text = as_text(raw_values[idx], fmt) if idx < len(raw_values) else None
            if canon in db_cols:
                vals[canon] = text
            elif text is not None:
                extras[str(header_values[idx])] = text
        if not any(vals.get(c) for c in columns):
            continue
        if extras:
            vals["extras"] = json.dumps(extras)
        rows_out.append(vals)
    return rows_out


def parse_simple_sheet(ws: Worksheet, spec: SimpleSheetSpec) -> list[dict[str, Any]]:
    rows_out = _extract_header_rows(ws, spec.header_row, spec.columns)
    if spec.literal_columns:
        for vals in rows_out:
            vals.update(spec.literal_columns)
    return rows_out


def parse_split_by_value_sheet(ws: Worksheet, spec: SplitByValueSheetSpec) -> dict[str, list[dict[str, Any]]]:
    rows = _extract_header_rows(ws, spec.header_row, spec.columns)
    results: dict[str, list[dict[str, Any]]] = {table: [] for table in spec.table_by_value.values()}
    for vals in rows:
        key = vals.get(spec.split_column)
        table = spec.table_by_value.get(str(key).strip().casefold()) if key else None
        if table:
            results[table].append(vals)
    return results


def _parse_control_center(ws: Worksheet) -> list[dict[str, Any]]:
    max_row = ws.max_row
    rows_out: list[dict[str, Any]] = []
    section: Optional[str] = None
    category: Optional[str] = None
    col3_role = "monthly_fee"

    for r in range(1, max_row + 1):
        a, b, c, d, e = _row_values(ws, r, 5)
        if _is_blank_row([a, b, c, d, e]):
            continue
        if a is not None and str(a).strip().casefold() == "category":
            section = as_text(b)
            col3_role = "fee_type" if d is not None and "fee type" in str(d).casefold() else "monthly_fee"
            category = None
            continue
        if c is None and b is not None:
            continue  # footnote / disclaimer row (wide-merged text, no service id)
        if a is not None:
            category = as_text(a)
        vals: dict[str, Any] = {
            "pricebook_ingestion_run_id": None,
            "excel_row_number": r,
            "section": section,
            "category": category,
            "item_description": as_text(b),
            "service_id": as_text(c),
            "monthly_fee": as_text(d, _cell_format(ws, r, 4)) if col3_role == "monthly_fee" else None,
            "overage_charges": as_text(e, _cell_format(ws, r, 5)),
            "fee_type": as_text(d) if col3_role == "fee_type" else None,
            "extras": None,
        }
        if any(vals.get(k) for k in ("item_description", "service_id", "monthly_fee", "fee_type")):
            rows_out.append(vals)
    return rows_out


def _parse_fleet_complete(ws: Worksheet) -> list[dict[str, Any]]:
    max_row = ws.max_row
    rows_out: list[dict[str, Any]] = []
    section: Optional[str] = None

    for r in range(1, max_row + 1):
        a, b, c, d = _row_values(ws, r, 4)
        if _is_blank_row([a, b, c, d]):
            continue
        if b is not None and str(b).strip().casefold() in ("high level overview of the plan", "description") and c is not None:
            section = as_text(a)
            continue
        if b is None and c is None and d is None:
            continue  # lone title row, e.g. "Fleet Complete Service Solution"
        vals = {
            "pricebook_ingestion_run_id": None,
            "excel_row_number": r,
            "section": section,
            "item": as_text(a),
            "description": as_text(b),
            "code": as_text(c),
            "price": as_text(d, _cell_format(ws, r, 4)),
            "extras": None,
        }
        rows_out.append(vals)
    return rows_out


def _parse_connected_worker(ws: Worksheet) -> list[dict[str, Any]]:
    max_row = ws.max_row
    rows_out: list[dict[str, Any]] = []
    pending_label: Optional[str] = None
    section: Optional[str] = None

    for r in range(1, max_row + 1):
        a, b, c, d, e = _row_values(ws, r, 5)
        if _is_blank_row([a, b, c, d, e]):
            continue
        if a is not None and str(a).strip().casefold() == "item" and b is not None and str(b).strip().casefold() == "description":
            section = pending_label
            continue
        if b is None and c is None and d is None and e is None and a is not None:
            pending_label = as_text(a)
            continue
        vals = {
            "pricebook_ingestion_run_id": None,
            "excel_row_number": r,
            "section": section,
            "item": as_text(a),
            "description": as_text(b),
            "service_id": as_text(c),
            "monthly_fee": as_text(d, _cell_format(ws, r, 4)),
            "dependencies": as_text(e),
            "extras": None,
        }
        rows_out.append(vals)
    return rows_out


def _parse_voice_data_usage_rates(ws: Worksheet) -> list[dict[str, Any]]:
    """Six repeating header blocks, e.g. ('Service ID'|'Parent Service ID',
    'Service', 'Description', 'Usage Rate'|'CPM Rate'). id_type/rate_type
    record which header variant each row's block used."""
    max_row = ws.max_row
    rows_out: list[dict[str, Any]] = []
    id_type: Optional[str] = None
    rate_type: Optional[str] = None

    for r in range(1, max_row + 1):
        a, b, c, d = _row_values(ws, r, 4)
        if _is_blank_row([a, b, c, d]):
            continue
        if b is not None and str(b).strip().casefold() == "service" and c is not None and str(c).strip().casefold() == "description":
            id_type = as_text(a)
            rate_type = as_text(d)
            continue
        vals = {
            "pricebook_ingestion_run_id": None,
            "excel_row_number": r,
            "id_type": id_type,
            "service_id": as_text(a),
            "service": as_text(b),
            "description": as_text(c),
            "rate_type": rate_type,
            "rate": as_text(d, _cell_format(ws, r, 4)),
            "extras": None,
        }
        rows_out.append(vals)
    return rows_out


def _parse_tls_sipa_v1(ws: Worksheet) -> list[dict[str, Any]]:
    """Header is row 2; data starts row 3. Column D ("Service ID (Monthly
    Top Up)") holds a second, real service ID on 5 of 15 rows — unpivoted
    into a second row (id_type='monthly_top_up') so both IDs are
    independently queryable. See catalogues.py for why monthly_fee is left
    NULL on the top-up row rather than copied or backfilled from
    overage_charges."""
    max_row = ws.max_row
    rows_out: list[dict[str, Any]] = []

    for r in range(3, max_row + 1):
        category, name, service_id, topup_id, desc, monthly_fee, overage = _row_values(ws, r, 7)
        if category is None and service_id is None:
            continue  # blank row, or a footnote row like "* As per specific M2M Shared Add-on"

        base_row = {
            "pricebook_ingestion_run_id": None,
            "excel_row_number": r,
            "product": "SIPA V1",
            "service_category": as_text(category),
            "service_name": as_text(name),
            "service_id": as_text(service_id),
            "id_type": "base",
            "short_service_description": as_text(desc),
            "monthly_fee": as_text(monthly_fee, _cell_format(ws, r, 6)),
            "overage_charges": as_text(overage, _cell_format(ws, r, 7)),
            "extras": None,
        }
        rows_out.append(base_row)

        topup_text = as_text(topup_id)
        if topup_text and topup_text.strip().casefold() != "n/a":
            topup_row = dict(base_row)
            topup_row["service_id"] = topup_text
            topup_row["id_type"] = "monthly_top_up"
            topup_row["monthly_fee"] = None
            rows_out.append(topup_row)
    return rows_out


MULTI_BLOCK_PARSERS: dict[str, Callable[[Worksheet], list[dict[str, Any]]]] = {
    "control_center": _parse_control_center,
    "fleet_complete": _parse_fleet_complete,
    "connected_worker": _parse_connected_worker,
    "voice_data_usage_rates": _parse_voice_data_usage_rates,
    "tls_sipa_v1": _parse_tls_sipa_v1,
}


def parse_workbook(path: Path, book: BookSpec) -> dict[str, list[dict[str, Any]]]:
    """Returns {table_name: rows} for every sheet in `book` found in the workbook.

    Raises if a sheet the book spec expects is missing from the file (a
    structural change worth failing loudly on, since this is a raw landing
    layer with no other validation).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    found_sheet_names = {name.strip().casefold() for name in wb.sheetnames}
    missing = [
        spec.sheet_name for spec in book.sheets if spec.sheet_name.strip().casefold() not in found_sheet_names
    ]
    if missing:
        raise ValueError(f"{path.name}: workbook is missing expected sheet(s): {', '.join(missing)}")

    results: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        spec = resolve_sheet(book, sheet_name)
        if spec is None:
            continue
        ws = wb[sheet_name]
        if isinstance(spec, SplitByValueSheetSpec):
            for table_name, rows in parse_split_by_value_sheet(ws, spec).items():
                results.setdefault(table_name, []).extend(rows)
            continue
        if isinstance(spec, MultiBlockSheetSpec):
            parser = MULTI_BLOCK_PARSERS[spec.parser]
            rows = parser(ws)
        else:
            rows = parse_simple_sheet(ws, spec)
        results.setdefault(spec.table_name, []).extend(rows)
    return results
