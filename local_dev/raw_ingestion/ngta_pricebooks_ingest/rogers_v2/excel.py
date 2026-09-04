"""Parse Rogers pricebook v2 Excel workbooks into per-table row dicts.

Both v2 books are plain header-row-plus-data-rows sheets, so this is the
header-canonicalization + row-extraction half of telus_v2/excel.py with the
merged-cell fill-down and multi-block/split-by-value machinery dropped —
neither book needs it (see catalogues.py).
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from rogers_v2.catalogues import BookSpec, SimpleSheetSpec, resolve_sheet


def canonical_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower().startswith("unnamed"):
        return ""
    s = re.sub(r"\s+", " ", s.lower())
    s = s.replace("(", " ").replace(")", " ").replace("*", " ")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def as_text(value: Any, number_format: Optional[str] = None) -> Optional[str]:
    """Stringify a cell value. Numeric cells with a currency or percentage
    number format (e.g. Rogers' `$0.00`-formatted Monthly Fixed Fee / Price
    columns, or Data Services' `0%`-formatted ECF Rate) are rendered that
    way rather than as the bare underlying number (0, 0.25) — a masked price
    still reads as "$0.00" text and a rate as "25%" text, consistent with
    the old PDF-parsed rogers/ tables (which read the rendered text
    directly) and with how telus_v2 handles the same situation."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool) and number_format:
        if "$" in number_format:
            return f"${value:,.2f}"
        if "%" in number_format:
            return f"{value * 100:g}%"
    s = str(value).strip()
    return s if s else None


def _row_values(ws: Worksheet, row: int, max_col: int) -> list[Any]:
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def _row_formats(ws: Worksheet, row: int, max_col: int) -> list[Optional[str]]:
    return [ws.cell(row=row, column=c).number_format for c in range(1, max_col + 1)]


def _is_blank_row(values: list[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def parse_simple_sheet(ws: Worksheet, spec: SimpleSheetSpec) -> list[dict[str, Any]]:
    max_col = ws.max_column
    max_row = ws.max_row

    header_values = _row_values(ws, spec.header_row, max_col)
    colmap: dict[int, str] = {}
    for idx, raw in enumerate(header_values):
        canon = canonical_header(raw)
        if canon:
            colmap[idx] = canon

    db_cols = frozenset(spec.columns)
    rows_out: list[dict[str, Any]] = []
    for r in range(spec.header_row + 1, max_row + 1):
        raw_values = _row_values(ws, r, max_col)
        if _is_blank_row(raw_values):
            continue
        raw_formats = _row_formats(ws, r, max_col)
        vals: dict[str, Any] = {"pricebook_ingestion_run_id": None, "excel_row_number": r, "extras": None}
        extras: dict[str, Any] = {}
        for idx, canon in colmap.items():
            fmt = raw_formats[idx] if idx < len(raw_formats) else None
            text = as_text(raw_values[idx], fmt) if idx < len(raw_values) else None
            if canon in db_cols:
                vals[canon] = text
            elif text is not None:
                extras[str(header_values[idx])] = text
        if not any(vals.get(c) for c in spec.columns):
            continue
        if extras:
            vals["extras"] = json.dumps(extras)
        if spec.literal_columns:
            vals.update(spec.literal_columns)
        rows_out.append(vals)
    return rows_out


def parse_workbook(path: Path, book: BookSpec) -> dict[str, list[dict[str, Any]]]:
    """Returns {table_name: rows} for every sheet in `book` found in the workbook.

    Raises if a sheet the book spec expects is missing from the file (a
    structural change worth failing loudly on, since this is a raw landing
    layer with no other validation).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    found_sheet_names = {name.strip().casefold() for name in wb.sheetnames}
    missing = [
        spec.sheet_name for spec in book.sheets if spec.sheet_name.strip().casefold() not in found_sheet_names
    ]
    if missing:
        raise ValueError(f"{path.name}: workbook is missing expected sheet(s): {', '.join(missing)}")

    results: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        spec = resolve_sheet(book, sheet_name)
        if spec is None:
            continue
        rows = parse_simple_sheet(wb[sheet_name], spec)
        results.setdefault(spec.table_name, []).extend(rows)
    return results
