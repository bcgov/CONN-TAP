import pandas as pd
import os
import glob
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

#----------FOLDER & FILES---------------------------------------------------------------------------------------------        


base_path = r'C:/Users/AFATHALI/OneDrive - Government of BC/Documents/Rogers_NGTA_Script/'
data_path = os.path.join(base_path, "Data")

ngta_files = glob.glob(os.path.join(data_path, "*NGTA - Administrator Reports_Usage_&_Spend*.xlsx"))
wireline_files = glob.glob(os.path.join(data_path, "BC_GOV_Wireline_report_*_Consolidated*.xlsx"))


#---------BGE MAPPING (handles ALL cases, including critical fixes for missing BGEs and school districts)-------------


BGE_MAP = {
    "BRITISH COLUMBIA LOTTERY CORPORATION": "BCLC",
    "BRITISH COLOMBIA LOTTERY CORPORATION": "BCLC",
    "BC LOTTERY CORPORATION": "BCLC",
    "BC LOTTERY": "BCLC",
    "BC HYDRO": "BC Hydro",
    "BRITISH COLUMBIA HYDRO": "BC Hydro",
    "EDUCATION AND CHILD CARE": "ECC",
    "FRASER HEALTH AUTHORITY": "FHA",
    "INTERIOR HEALTH AUTHORITY": "IHA",
    "NORTHERN HEALTH AUTHORITY": "NHA",
    "INSURANCE CORPORATION OF BRITISH COLUMB.": "ICBC",
    "PROVINCIAL HEALTH SERVICES AUTHORITY": "PHSA",
    "VANCOUVER COASTAL HEALTH AUTHORITY": "VCHA (+PHC)",
    "PROVIDENCE HEALTH CARE": "VCHA (+PHC)",
    "VANCOUVER ISLAND HEALTH AUTHORITY": "VIHA",
    "BC GOVERNMENT MINISTRIES": "Gov BC",
}

SUB_BGE_TO_BGE = {
    "BC MIN ATTORNEY GENERAL": "Gov BC",
    "INDIGENOUS RELATIONS AND RECONCILIATION": "Gov BC",
    "MINISTRY OF CITIZENS SERVICES": "Gov BC",
    "MINISTRY OF INFRASTRUCTURE": "Gov BC",
    "MIN OF SOCIAL DEVELOPMENT": "Gov BC",
}


#--------DATE EXTRACTION (handles both full and partial date formats in filenames)---------------------------------


def extract_month(filename):
    match_full = re.search(r'(\d{8})', filename)
    if match_full:
        return pd.to_datetime(match_full.group(1), format="%Y%m%d").strftime("%b %Y")

    match_partial = re.search(r'(\d{4})[_-](\d{2})', filename)
    if match_partial:
        y, m = match_partial.groups()
        return pd.to_datetime(f"{y}-{m}-01").strftime("%b %Y")

    return None


# --CRITICAL FIX (handles ALL cases)
def resolve_bge(row):
    bge = str(row.get('BGE', '')).strip().upper()
    sub = str(row.get('SUB-BGE', '')).strip().upper()

    # SCHOOL DISTRICT
    if sub.startswith("SCHOOL DISTRICT") or sub.startswith("DISTRICT"):
        return "School Districts"

    # FORCE FIX (robust match)
    if "FAMILY MAINTENANCE AGENCY" in sub or "PUBLIC SERVICE AGENCY" in sub:
        return "Gov BC"

    val = bge if bge else sub
    val = SUB_BGE_TO_BGE.get(val, val)
    val = BGE_MAP.get(val, val)

    return val



# --------LOAD ROGERS NGTA -------------------------------------------------------------------------------------


all_data = []

for file in ngta_files:
    print(f"\nProcessing NGTA: {file}")

    df = pd.read_excel(file, engine='openpyxl')
    df.columns = df.columns.str.strip()

    if 'BGE' not in df.columns and 'SUB-BGE' not in df.columns:
        continue

    df['BGE'] = df.apply(resolve_bge, axis=1)

    billed_col = 'BILLED_AMOUNT(PRE-TAX)' if 'BILLED_AMOUNT(PRE-TAX)' in df.columns else 'CHARGES_SUBTOTAL'

    month = extract_month(os.path.basename(file))
    if not month:
        continue

    grouped = df.groupby('BGE').agg({
        billed_col: 'sum',
        'HARDWARE': 'sum'
    }).reset_index()

    grouped.rename(columns={billed_col: 'BILLED'}, inplace=True)
    grouped['CELLULAR PLAN'] = grouped['BILLED'] - grouped['HARDWARE']
    grouped['Month'] = month

    all_data.append(grouped[['BGE', 'Month', 'HARDWARE', 'CELLULAR PLAN']])

final_df = pd.concat(all_data, ignore_index=True)


# --------LOAD WIRELINE REPORT --------------------------------------------------------------------------------



wireline_data = []

for file in wireline_files:
    print(f"\nWireline: {file}")
    df = pd.read_excel(file, engine='openpyxl')
    df.columns = df.columns.str.strip().str.upper()

    if 'BGE' not in df.columns:
        continue

    df['BGE'] = df['BGE'].apply(lambda x: "School District"
                               if str(x).upper().startswith("SCHOOL DISTRICT")
                               else BGE_MAP.get(str(x).upper(), str(x).upper()))

    df['CATEGORY'] = df['PRODUCTLINE'].apply(
        lambda x: 'VOICE' if str(x).upper() == 'VOICE'
        else 'DATA' if str(x).upper() == 'DATA'
        else 'OTHER'
    )

    month = extract_month(os.path.basename(file))
    if not month:
        continue

    grouped = df.groupby(['BGE', 'CATEGORY'])['BILLED_AMOUNT(PRE-TAX)'].sum().reset_index()
    grouped['Month'] = month

    wireline_data.append(grouped)

wireline_df = pd.concat(wireline_data, ignore_index=True)



# -------MONTHS ORDER (full range from Jan 2024 to Dec 2026, to ensure all months are included even if missing in data)------


months = pd.date_range("2024-01-01", "2026-12-01", freq='MS')
month_order = [m.strftime("%b %Y") for m in months]



# -----------BUILD REPORT---------------------------------------------------------------------------------------------------


rows = []

total_df = final_df.groupby('Month').sum(numeric_only=True).reset_index()
wire_total = wireline_df.groupby(['Month', 'CATEGORY'])['BILLED_AMOUNT(PRE-TAX)'].sum().reset_index()

# Rogers NGTA totals (overall, not by BGE)
for metric in ['HARDWARE', 'CELLULAR PLAN']:
    row = {'Section': 'Rogers NGTA', 'Metric': metric}
    for m in month_order:
        val = total_df.loc[total_df['Month'] == m, metric]
        row[m] = val.values[0] if not val.empty else ""
    rows.append(row)

for cat in ['VOICE', 'DATA', 'OTHER']:
    row = {'Section': 'Rogers NGTA', 'Metric': cat}
    for m in month_order:
        val = wire_total[(wire_total['Month'] == m) & (wire_total['CATEGORY'] == cat)]
        row[m] = val['BILLED_AMOUNT(PRE-TAX)'].values[0] if not val.empty else ""
    rows.append(row)

row = {'Section': 'Rogers NGTA', 'Metric': 'Total'}
for m in month_order:
    ngta = total_df.loc[total_df['Month'] == m, ['HARDWARE', 'CELLULAR PLAN']].sum(axis=1)
    wire = wire_total.loc[wire_total['Month'] == m, 'BILLED_AMOUNT(PRE-TAX)'].sum()
    row[m] = (ngta.values[0] if not ngta.empty else 0) + wire
rows.append(row)

# BGEs Section     
required_bges = ["ECC"]

all_bges = sorted(set(final_df['BGE'].unique()).union(required_bges))

for bge in all_bges:
    bge_df = final_df[final_df['BGE'] == bge]
    grouped = bge_df.groupby('Month').sum(numeric_only=True).reset_index()

    for metric in ['HARDWARE', 'CELLULAR PLAN']:
        row = {'Section': bge, 'Metric': metric}
        for m in month_order:
            val = grouped.loc[grouped['Month'] == m, metric]
            row[m] = val.values[0] if not val.empty else ""
        rows.append(row)

    for cat in ['VOICE', 'DATA', 'OTHER']:
        row = {'Section': bge, 'Metric': cat}
        for m in month_order:
            val = wireline_df[
                (wireline_df['Month'] == m) &
                (wireline_df['CATEGORY'] == cat) &
                (wireline_df['BGE'] == bge)
            ]
            row[m] = val['BILLED_AMOUNT(PRE-TAX)'].sum() if not val.empty else ""
        rows.append(row)

    row = {'Section': bge, 'Metric': 'Total'}
    for m in month_order:
        h = grouped.loc[grouped['Month'] == m, 'HARDWARE']
        c = grouped.loc[grouped['Month'] == m, 'CELLULAR PLAN']
        ngta = (h.values[0] if not h.empty else 0) + (c.values[0] if not c.empty else 0)

        wire = wireline_df[
            (wireline_df['Month'] == m) &
            (wireline_df['BGE'] == bge)
        ]['BILLED_AMOUNT(PRE-TAX)'].sum()

        row[m] = ngta + wire

    rows.append(row)

report_df = pd.DataFrame(rows)



# ----------------METRIC RENAME + ORDER----------------------------------------------------------------------

metric_map = {
    "CELLULAR PLAN": "Cellular Plans",
    "HARDWARE": "Cellular H/W",
    "VOICE": "Voice",
    "DATA": "Data",
    "OTHER": "Other",
    "Total": "Total"
}

report_df['Metric'] = report_df['Metric'].map(lambda x: metric_map.get(x, x))

sections = report_df['Section'].dropna().unique().tolist()
ordered_sections = ['Rogers NGTA'] + sorted([s for s in sections if s != 'Rogers NGTA'])

report_df['Section'] = pd.Categorical(report_df['Section'], categories=ordered_sections, ordered=True)

metric_order = ["Cellular Plans", "Cellular H/W", "Data", "Voice", "Other", "Total"]

report_df['Metric'] = pd.Categorical(report_df['Metric'], categories=metric_order, ordered=True)

report_df = report_df.sort_values(by=['Section', 'Metric'], kind='stable')

# ----------------ADD TOTAL ROWS (ALL BGEs EXCEPT ROGERS)-----------------------------------------------------


# Get only BGE rows (exclude Rogers)
bge_only = report_df[report_df['Section'] != 'Rogers NGTA'].copy()

# Remove blank section rows (after cleaning)
bge_only = bge_only[bge_only['Section'] != ""]

# Metrics to total
total_metrics_map = {
    "Cellular Plans": "TOTAL Cellular Plans",
    "Cellular H/W": "TOTAL Cellular H/W",
    "Data": "TOTAL Data",
    "Voice": "TOTAL Voice",
    "Other": "TOTAL Other"
}

total_rows = []

for metric, label in total_metrics_map.items():

    metric_df = bge_only[bge_only['Metric'] == metric]

    row = {
        'Section': "",   # <-- IMPORTANT (blank section)
        'Metric': label
    }

    for m in month_order:
        vals = pd.to_numeric(metric_df[m], errors='coerce')
        total = vals.sum(skipna=True)

        row[m] = total if total != 0 else 0  # keep 0 for BGE totals

    total_rows.append(row)

# Append to report
report_df = pd.concat([report_df, pd.DataFrame(total_rows)], ignore_index=True)

# ------CLEAN SECTION LABELS (SHOW ONLY FIRST ROW)--------------------------------------------------------------

last = None
rows_clean = []

for _, r in report_df.iterrows():
    if r['Section'] == last:
        r['Section'] = ""
    else:
        last = r['Section']
    rows_clean.append(r)

report_df = pd.DataFrame(rows_clean)



# SAVE OUTPUT (before formatting, to preserve data types for Excel formatting step)---------------------------------


output_file = os.path.join(base_path, 'rogers_spend_report.xlsx')
report_df.to_excel(output_file, index=False, header=False)



# -------------FORMAT EXCEL (FULL)--------------------------------------------------------------------------------


wb = load_workbook(output_file)
ws = wb.active

header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
rogers_fill = PatternFill(start_color="ED7D31", fill_type="solid")
bge_fill = PatternFill(start_color="F8CBAD", fill_type="solid")

ws.insert_rows(1)

for i, m in enumerate(month_order, start=3):
    cell = ws.cell(row=1, column=i, value=m)
    cell.fill = header_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

current = None
for r in range(2, ws.max_row + 1):

    sec_cell = ws.cell(row=r, column=1)
    metric_cell = ws.cell(row=r, column=2)

    if sec_cell.value not in (None, ""):
        current = sec_cell.value
        sec_cell.font = Font(bold=True)

    is_total_row = str(metric_cell.value).startswith("TOTAL")

    fill = rogers_fill if (current == "Rogers NGTA" or is_total_row) else bge_fill

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        if c != 1:
            cell.font = Font(color="000000")

    if metric_cell.value == "Total":
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).font = Font(bold=True)

# currency
for r in range(2, ws.max_row + 1):
    for c in range(3, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '"$"#,##0.00;("$"#,##0.00)'


# ROGERS_NGTA section: BLANK show as 0.00 

current_section = None

for r in range(2, ws.max_row + 1):
    sec = ws.cell(row=r, column=1).value

    if sec not in (None, ""):
        current_section = sec

    # Only apply to Rogers NGTA
    if current_section == "Rogers NGTA":
        for c in range(3, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)

            if cell.value in (None, ""):
                cell.value = 0
                cell.number_format = '"$"#,##0.00;("$"#,##0.00)'


# negative red
for r in range(2, ws.max_row + 1):
    for c in range(3, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, (int, float)) and cell.value < 0:
            cell.font = Font(color="FF0000", bold=cell.font.bold)

# auto width
for col in range(1, ws.max_column + 1):
    max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
    ws.column_dimensions[get_column_letter(col)].width = max_len + 3

# freeze (Rogers block)
freeze_row = None
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val not in (None, "", "Rogers NGTA"):
        freeze_row = r
        break

ws.freeze_panes = f"C{freeze_row}" if freeze_row else "C2"

# border between sections
thin = Side(style="thin")
border = Border(top=thin)

for r in range(3, ws.max_row + 1):
    curr = ws.cell(row=r, column=1).value
    prev = ws.cell(row=r-1, column=1).value
    if curr not in (None, "") and curr != prev:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border

wb.save(output_file)
