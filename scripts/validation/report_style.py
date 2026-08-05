"""Shared Excel presentation for the validation reports (TELUS + Rogers).

`style_workbook(path)` post-processes a written .xlsx purely cosmetically -- it does not
change any values or validation logic. Applied by every runner so all reports look the same:

  * bold white header row on a solid blue fill, frozen, with an autofilter
  * auto-sized column widths (sampled, so it stays fast on large sheets)
  * Summary tab: PASS/FAIL/SKIPPED status cells colored green/red/amber, and a tab colour

It is best-effort: callers wrap it in try/except so a styling hiccup never fails a report.
"""

from __future__ import annotations

from pathlib import Path

# Palette (hex, no leading '#').
HEADER_FILL = "305496"   # blue
HEADER_TEXT = "FFFFFF"   # white
PASS_FILL = "C6EFCE"     # green
FAIL_FILL = "FFC7CE"     # red
SKIP_FILL = "FFEB9C"     # amber
SUMMARY_TAB = "305496"   # blue tab for the Summary sheet

# Legend box on the New-Removed sheet -- deliberately different from the blue data table so it
# reads as a separate reference card.
LEGEND_HEADER_FILL = "44546A"   # slate
LEGEND_BODY_FILL = "F2F2F2"     # light grey
LEGEND_BORDER = "BFBFBF"        # grey
LEGEND_GAP = 3                  # columns from the data to the legend (2 blank + legend start)

# Column-width bounds and how many data rows to sample when measuring width.
MIN_WIDTH = 10
MAX_WIDTH = 60
WIDTH_SAMPLE_ROWS = 200

# Status legend written to the side of the New-Removed detection sheet. Rogers and TELUS share
# one status vocabulary -- the Rogers detection functions (rogers_*_new_removed_detection) were
# renamed to match the TELUS ones (telus_raw_validate_new_bges_in_sheets /
# telus_raw_validate_new_sub_bges_in_accounts) -- so one list covers both providers. The sheet
# gets the FULL set: every possible status, not just the ones in this month's data. Descriptions
# come from the detection functions, not assumptions.
NEW_REMOVED_TAB = "New-Removed BGEs"
STATUS_LEGEND = [
    ("Unmapped",
     "In the current month, absent in the prior month, and not a recognized alias in the seeds."),
    ("Persisting Unmapped",
     "In both the prior and current month, but still not a recognized alias in the seeds."),
    ("New Match",
     "In the current month, absent in the prior month, and a recognized alias in the seeds."),
    ("Disappeared",
     "In the prior month but not in the current month."),
    ("Still Disappeared",
     "Present two months ago, absent in both the prior and current month. Rogers only -- the "
     "TELUS detection compares against the prior month alone and has no two-month status."),
]


def _best_width(col_cells) -> int:
    longest = 0
    for cell in col_cells[: WIDTH_SAMPLE_ROWS + 1]:
        if cell.value is not None:
            longest = max(longest, len(str(cell.value)))
    return max(MIN_WIDTH, min(MAX_WIDTH, longest + 2))


def _add_status_legend(ws, header_font, header_align) -> None:
    """Write a styled Status/Description legend as a separate card to the right of the data.

    Documents every status the detection functions can emit -- not just the ones in this month's
    data -- so an absent status reads as "did not occur" rather than "not a thing".
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    entries = STATUS_LEGEND

    header_fill = PatternFill("solid", fgColor=LEGEND_HEADER_FILL)
    body_fill = PatternFill("solid", fgColor=LEGEND_BODY_FILL)
    edge = Side(style="thin", color=LEGEND_BORDER)
    border = Border(left=edge, right=edge, top=edge, bottom=edge)
    bold_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    start = ws.max_column + LEGEND_GAP  # blank gap columns between the data and the legend
    for offset, title in enumerate(("Status", "Description")):
        cell = ws.cell(row=1, column=start + offset, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
    for i, (status, desc) in enumerate(entries, start=2):
        status_cell = ws.cell(row=i, column=start, value=status)
        desc_cell = ws.cell(row=i, column=start + 1, value=desc)
        status_cell.font = bold_font
        for cell in (status_cell, desc_cell):
            cell.fill = body_fill
            cell.border = border
            cell.alignment = wrap
    ws.column_dimensions[get_column_letter(start)].width = 22
    ws.column_dimensions[get_column_letter(start + 1)].width = 80


def _color_summary_status(ws, status_fills) -> None:
    """Give the Summary tab a colour and shade its PASS/FAIL/SKIPPED status cells."""
    ws.sheet_properties.tabColor = SUMMARY_TAB
    status_col = {c.value: c.column for c in ws[1]}.get("Status")
    if not status_col:
        return
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        fill = status_fills.get(str(cell.value))
        if fill:
            cell.fill = fill


def _style_sheet(ws, header_fill, header_font, header_align, status_fills) -> None:
    """Apply header styling, freeze/filter, widths, and any per-tab extras to one sheet."""
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 18

    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = _best_width(col_cells)

    if ws.title == "Summary":
        _color_summary_status(ws, status_fills)
    elif ws.title == NEW_REMOVED_TAB:
        _add_status_legend(ws, header_font, header_align)


def style_workbook(path: str | Path) -> None:
    """Apply the shared cosmetic styling to every sheet of the workbook at `path`."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_TEXT)
    header_align = Alignment(vertical="center")
    status_fills = {
        "PASS": PatternFill("solid", fgColor=PASS_FILL),
        "FAIL": PatternFill("solid", fgColor=FAIL_FILL),
        "SKIPPED": PatternFill("solid", fgColor=SKIP_FILL),
    }

    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row >= 1 and ws.max_column >= 1:
            _style_sheet(ws, header_fill, header_font, header_align, status_fills)
    wb.save(path)
