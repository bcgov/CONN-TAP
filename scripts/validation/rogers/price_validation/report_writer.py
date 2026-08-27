# report_writer.py
# Excel-writing logic extracted verbatim from rogers_wireless_price_validation.py's
# "Create Excel workbook" / "Add Summary sheet" / "Save workbook" sections, wrapped
# in a function so main.py / RUNBOOK.ipynb can call the exact same code instead of
# re-implementing it.

import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


def write_price_validation_workbook(
    comparison,
    summary_rows,
    output_file,
    missing_service_id_df=None,
):
    """
    comparison: DataFrame containing every comparison row (Matched /
        Missing from Price Book / Missing Service ID from Report).
    summary_rows: list of (label, value) tuples for the Summary sheet.
    output_file: path to write the .xlsx to.
    missing_service_id_df: the missing_service_id_report DataFrame from
        rogers_wireless_price_validation.py (its own, wider column set). If
        omitted, the "Missing Service ID from Report" sheet is just
        comparison filtered to that status, using comparison's columns.
    """
    comparison_columns = list(comparison.columns)

    if missing_service_id_df is not None:
        missing_service_id_report = missing_service_id_df
    else:
        missing_service_id_report = comparison[
            comparison["Match Status (Service ID)"] == "Missing Service ID from Report"
        ]

    mismatched_rate = comparison[
        (
            comparison["Match Status (Service ID)"] == "Matched"
        )
        &
        (
            comparison["Difference (Billed Amount - Monthly Fixed Fee)"]
            .abs() > 0.005
        )
    ]

    # =========================
    # Create Excel workbook
    # =========================

    wb = Workbook()

    views = [
        ("Complete comparison", comparison, comparison_columns),
        ("Matched Service IDs", comparison[comparison["Match Status (Service ID)"] == "Matched"], comparison_columns),
        (
            "Missing from Price Book",
            comparison[comparison["Match Status (Service ID)"] == "Missing from Price Book"],
            comparison_columns,
        ),
        (
            "Missing Service ID from Report",
            missing_service_id_report,
            list(missing_service_id_report.columns),
        ),
        (
            "Mismatched Rate",
            mismatched_rate,
            comparison_columns,
        ),

    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    matched_fill = PatternFill("solid", fgColor="E2F0D9")
    missing_fill = PatternFill("solid", fgColor="FCE4D6")
    missing_service_fill = PatternFill(
        "solid",
        fgColor="FFF2CC"
    )
    difference_fill = PatternFill("solid", fgColor="FFF2CC")

    used_table_names = set()

    for sheet_index, (sheet_name, df, sheet_columns) in enumerate(views):
        if sheet_index == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet(sheet_name[:31])

        ws.title = sheet_name[:31]
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        # Headers
        for col_index, header in enumerate(sheet_columns, start=1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Dynamic column positions for comparison sheets
        status_col_index = None
        monthly_fee_col_letter = None
        billed_rate_col_letter = None

        if "Match Status (Service ID)" in sheet_columns:
            status_col_index = sheet_columns.index("Match Status (Service ID)") + 1

        if "Monthly Fixed Fee (from the Price Book)" in sheet_columns:
            monthly_fee_col_letter = get_column_letter(
                sheet_columns.index("Monthly Fixed Fee (from the Price Book)") + 1
            )

        if "Billed Amount (Rate)" in sheet_columns:
            billed_rate_col_letter = get_column_letter(
                sheet_columns.index("Billed Amount (Rate)") + 1
            )

        # Data rows
        for row_index, row in enumerate(df.itertuples(index=False), start=2):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="top",
                    wrap_text=True,
                )

                column_name = sheet_columns[col_index - 1]

                if column_name in [
                    "Monthly Fixed Fee (from the Price Book)",
                    "Billed Amount (Rate)",
                    "Billed Amount Numeric",
                ]:
                    cell.number_format = '$#,##0.00;[Red]-$#,##0.00;-'

                if column_name == "Difference (Billed Amount - Monthly Fixed Fee)":
                    if monthly_fee_col_letter and billed_rate_col_letter:
                        cell.value = (
                            f'=IF(OR({monthly_fee_col_letter}{row_index}="",'
                            f'{billed_rate_col_letter}{row_index}=""),"",'
                            f'{billed_rate_col_letter}{row_index}-{monthly_fee_col_letter}{row_index})'
                        )
                    cell.number_format = '$#,##0.00;$#,##0.00;-'

            # Status highlighting, if this sheet has Match Status
            if status_col_index:
                status_cell = ws.cell(row=row_index, column=status_col_index)
                status = status_cell.value

                if status == "Matched":
                  status_cell.fill = matched_fill

                elif status == "Missing from Price Book":
                  status_cell.fill = missing_fill

                elif status == "Missing Service ID from Report":
                   status_cell.fill = missing_service_fill

        # Column widths
        for col_index, header in enumerate(sheet_columns, start=1):
            width = max(14, min(42, len(str(header)) + 4))
            ws.column_dimensions[get_column_letter(col_index)].width = width

        # Override widths for comparison sheets
        if sheet_name != "Missing Service ID from Report":
            widths = {
                1: 26,
                2: 30,
                3: 24,
                4: 32,
                5: 24,
                6: 24,
                7: 22,
                8: 18,
                9: 38,
            }

            for col_index, width in widths.items():
                if col_index <= len(sheet_columns):
                    ws.column_dimensions[get_column_letter(col_index)].width = width

        ws.row_dimensions[1].height = 36

        # Add Excel table
        max_row = max(1, len(df) + 1)
        max_col = len(sheet_columns)
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

        if max_row >= 2 and max_col >= 1:
            base_table_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name[:20]).strip("_")

            if not base_table_name:
                base_table_name = f"Table{sheet_index + 1}"

            table_name = base_table_name
            suffix = 1
            while table_name in used_table_names:
                suffix += 1
                table_name = f"{base_table_name}_{suffix}"

            used_table_names.add(table_name)

            table = Table(displayName=table_name, ref=table_ref)

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            table.tableStyleInfo = style
            ws.add_table(table)

    # Highlight non-zero differences only on sheets that have the Difference column
        if "Difference (Billed Amount - Monthly Fixed Fee)" in sheet_columns and max_row >= 2:

          difference_col = get_column_letter(
            sheet_columns.index(
                "Difference (Billed Amount - Monthly Fixed Fee)"
            ) + 1
          )

          red_fill = PatternFill(
            "solid",
            fgColor="FFC7CE"
          )

          red_font = Font(
            color="E06666",
            bold=True
          )

          ws.conditional_formatting.add(
            f"{difference_col}2:{difference_col}{max_row}",
            FormulaRule(
                formula=[
                    f'AND(ISNUMBER({difference_col}2),ABS({difference_col}2)>0.005)'
                ],
                fill=red_fill,
                font=red_font,
            ),
          )

    # =========================
    # Add Summary sheet
    # =========================

    summary = wb.create_sheet("Summary")
    summary.sheet_view.showGridLines = False

    summary["A1"] = "Summary"
    summary["A1"].font = Font(bold=True, size=14, color="1F4E78")

    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row_index, column=2, value=value)

    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 18


    # =========================
    # Save workbook
    # =========================

    wb.save(output_file)
