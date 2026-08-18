import pandas as pd
import re
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Font
# =========================
# File paths
# =========================

DATA_DIR = "."

price_xlsx = os.path.join(DATA_DIR, "cellular.xlsx")
report_xlsx = os.path.join(
    DATA_DIR,
    "2026_06_BC Gov't NGTA - Administrator Reports_Usage_&_Spend.xlsx"
)
output_file = os.path.join(DATA_DIR, "Wireless_Service_ID_Comparison_Rate.xlsx")


# =========================
# Load source files
# =========================

price = pd.read_excel(price_xlsx, engine="openpyxl")
report = pd.read_excel(
    report_xlsx,
    dtype=str,
    engine="openpyxl"
)

# Clean column names
price.columns = [str(c).strip() for c in price.columns]
report.columns = [str(c).strip() for c in report.columns]


# =========================
# Helper functions
# =========================

def find_col(columns, options):
    """
    Robustly find a column by comparing simplified column names.
    """
    normalized = {
        re.sub(r"[^a-z0-9]", "", c.lower()): c
        for c in columns
    }

    for option in options:
        key = re.sub(r"[^a-z0-9]", "", option.lower())
        if key in normalized:
            return normalized[key]

    for c in columns:
        c_lower = c.lower()
        if any(option.lower() in c_lower for option in options):
            return c

    raise KeyError(f"Missing column among: {options}")


def normalize_service_id(value):
    """
    Normalize Service IDs.

    Example:
    CVDULTDM -> CVDULTD
    """
    if pd.isna(value):
        return ""

    service_id = str(value).strip()

    # Normalize monthly variant IDs, e.g. CVDULTDM -> CVDULTD
    if service_id.endswith("M") and len(service_id) > 1:
        return service_id[:-1]

    return service_id


def parse_money(value):
    """
    Extract numeric amount from fee text.

    Examples:
    "$7.49 per seat" -> 7.49
    "$12.00" -> 12.00
    "No Charge" -> 0.00
    "($10.00)" -> -10.00
    "-10.00" -> -10.00
    """
    if pd.isna(value):
        return None

    text = str(value).strip()

    if not text:
        return None

    lower_text = text.lower()

    if "no charge" in lower_text:
        return 0.0

    if lower_text in {"n/a", "na", "none"}:
        return None

    # Remove commas and dollar signs from values like $1,996.00
    text = text.replace(",", "").replace("$", "")

    # Handle accounting-style negatives: ($12.34)
    is_parentheses_negative = text.startswith("(") and text.endswith(")")
    text = text.replace("(", "").replace(")", "")

    match = re.search(r"[-+]?\s*([0-9]+(?:\.[0-9]+)?)", text)

    if match:
        amount = float(match.group(1))
        if text.strip().startswith("-") or is_parentheses_negative:
            amount = -amount
        return amount

    return None


# =========================
# Identify required columns
# =========================

price_service_id_col = find_col(price.columns, ["Service ID"])
price_fee_col = find_col(price.columns, ["Monthly Fixed Fee"])

report_service_id_col = find_col(report.columns, ["SERVICE_ID", "Service ID"])
report_billed_col = find_col(
    report.columns,
    ["BILLED_AMOUNT(PRE-TAX)", "BILLED_AMOUNT(PRE TAX)", "Billed Amount Pre Tax"]
)


# =========================
# Prepare Price Book data
# =========================

price["Normalized Service ID"] = price[price_service_id_col].map(normalize_service_id)
price["Monthly Fixed Fee Numeric"] = price[price_fee_col].map(parse_money)

price_clean = price[
    price["Normalized Service ID"].astype(str).str.strip().ne("")
].copy()

price_clean = (
    price_clean[
        [
            price_service_id_col,
            "Normalized Service ID",
            price_fee_col,
            "Monthly Fixed Fee Numeric",
        ]
    ]
    .drop_duplicates("Normalized Service ID", keep="first")
)


# =========================
# Prepare Monthly Report data
# =========================

report["Billed Amount Numeric"] = report[report_billed_col].map(parse_money)

# Remove all report rows where BILLED AMOUNT(PRE TAX) is exactly zero.
# Rows with blank or non-numeric billed amount are kept for review.
zero_billed_count = int(report["Billed Amount Numeric"].eq(0).sum())
report_nonzero = report[report["Billed Amount Numeric"].ne(0)].copy()

# Capture report rows with blank SERVICE_ID after removing zero-billed rows.
blank_service_id_mask = (
    report_nonzero[report_service_id_col].isna()
    | report_nonzero[report_service_id_col].astype(str).str.strip().eq("")
)

missing_service_id_report = report_nonzero[blank_service_id_mask].copy()

# Put helper columns near the front for easier review in the Missing Service ID sheet.
missing_service_id_front_cols = [report_service_id_col, report_billed_col, "Billed Amount Numeric"]
missing_service_id_columns = (
    missing_service_id_front_cols
    + [c for c in missing_service_id_report.columns if c not in missing_service_id_front_cols]
)
missing_service_id_report = missing_service_id_report[missing_service_id_columns]

# Exclude blank SERVICE_ID rows from price-book comparison.
report_for_comparison = report_nonzero[~blank_service_id_mask].copy()

report_for_comparison["Normalized Service ID"] = report_for_comparison[
    report_service_id_col
].map(normalize_service_id)

# Keep ONLY required monthly report columns for comparison.
monthly_report_clean = report_for_comparison[
    [
        report_service_id_col,
        "Normalized Service ID",
        "Billed Amount Numeric",
        "MSF_OTHER_OPTIONS",
        "HARDWARE",
        "OTHERS",
    ]
].copy()

monthly_report_clean.rename(
    columns={
        report_service_id_col: "Monthly Report Service ID",
        "Billed Amount Numeric": "Billed Amount (Rate)",
    },
    inplace=True
)


# =========================
# Build comparison table
# =========================

comparison = monthly_report_clean.merge(
    price_clean[
        [
            price_service_id_col,
            "Normalized Service ID",
            "Monthly Fixed Fee Numeric",
        ]
    ],
    on="Normalized Service ID",
    how="left"
)

comparison.rename(
    columns={
        price_service_id_col: "Price Book Service ID",
        "Monthly Fixed Fee Numeric": "Monthly Fixed Fee (from the Price Book)",
    },
    inplace=True
)

# Add Match Status
comparison["Match Status (Service ID)"] = comparison["Price Book Service ID"].apply(
    lambda x: "Matched" if pd.notna(x) else "Missing from Price Book"
)

# Add difference column. The Excel formula is added later when writing each row.
comparison["Difference (Billed Amount - Monthly Fixed Fee)"] = (
    comparison["Billed Amount (Rate)"]
    - comparison["Monthly Fixed Fee (from the Price Book)"]
)

# Sort output
status_order = {
    "Missing Service ID from Report": 0,
    "Missing from Price Book": 2,
    "Matched": 1,
}
comparison["__sort_order"] = (
    comparison["Match Status (Service ID)"]
    .map(status_order)
    .fillna(9)
)

comparison = (
    comparison
    .sort_values("__sort_order")
    .drop(columns="__sort_order")
)

comparison_desired_order = [
    "Price Book Service ID",
    "Monthly Report Service ID",
    "Match Status (Service ID)",
    "Monthly Fixed Fee (from the Price Book)",
    "Billed Amount (Rate)",
    "MSF_OTHER_OPTIONS",
    "HARDWARE",
    "OTHERS",
    "Difference (Billed Amount - Monthly Fixed Fee)",
]

comparison_columns = [
    col for col in comparison_desired_order
    if col in comparison.columns
]

comparison = comparison[comparison_columns]

mismatched_rate = comparison[
    (
        comparison["Match Status (Service ID)"] == "Matched"
    )
    &
    (
        comparison["Difference (Billed Amount - Monthly Fixed Fee)"]
        .abs() > 0.005
    )
].copy()

# Add Missing Service ID rows to Complete Comparison

missing_service_comparison = pd.DataFrame({
    "Price Book Service ID": None,
    "Monthly Report Service ID": missing_service_id_report[report_service_id_col],
    "Match Status (Service ID)": "Missing Service ID from Report",
    "Monthly Fixed Fee (from the Price Book)": None,
    "Billed Amount (Rate)": missing_service_id_report["Billed Amount Numeric"],
    "MSF_OTHER_OPTIONS": missing_service_id_report["MSF_OTHER_OPTIONS"],
    "HARDWARE": missing_service_id_report["HARDWARE"],
    "OTHERS": missing_service_id_report["OTHERS"],
    "Difference (Billed Amount - Monthly Fixed Fee)": None
})
comparison = pd.concat(
    [
        comparison,
        missing_service_comparison
    ],
    ignore_index=True
)


# =========================
# Create Excel workbook
# =========================

wb = Workbook()

views = [
    ("Complete comparison", comparison, comparison_columns),
    ("Matched Service IDs", comparison[comparison["Match Status (Service ID)"] == "Matched"], comparison_columns),
    (
        "Missing from Price Book",
        comparison[comparison["Match Status (Service ID)"] == "Missing from Price Book"],
        comparison_columns,
    ),
    (
        "Missing Service ID from Report",
        missing_service_id_report,
        list(missing_service_id_report.columns),
    ),
    (
        "Mismatched Rate",
        mismatched_rate,
        comparison_columns,
    ),

]

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

matched_fill = PatternFill("solid", fgColor="E2F0D9")
missing_fill = PatternFill("solid", fgColor="FCE4D6")
missing_service_fill = PatternFill(
    "solid",
    fgColor="FFF2CC"
)
difference_fill = PatternFill("solid", fgColor="FFF2CC")

used_table_names = set()

for sheet_index, (sheet_name, df, sheet_columns) in enumerate(views):
    if sheet_index == 0:
        ws = wb.active
    else:
        ws = wb.create_sheet(sheet_name[:31])

    ws.title = sheet_name[:31]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Headers
    for col_index, header in enumerate(sheet_columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # Dynamic column positions for comparison sheets
    status_col_index = None
    monthly_fee_col_letter = None
    billed_rate_col_letter = None

    if "Match Status (Service ID)" in sheet_columns:
        status_col_index = sheet_columns.index("Match Status (Service ID)") + 1

    if "Monthly Fixed Fee (from the Price Book)" in sheet_columns:
        monthly_fee_col_letter = get_column_letter(
            sheet_columns.index("Monthly Fixed Fee (from the Price Book)") + 1
        )

    if "Billed Amount (Rate)" in sheet_columns:
        billed_rate_col_letter = get_column_letter(
            sheet_columns.index("Billed Amount (Rate)") + 1
        )

    # Data rows
    for row_index, row in enumerate(df.itertuples(index=False), start=2):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(
                horizontal="right",
                vertical="top",
                wrap_text=True,
            )

            column_name = sheet_columns[col_index - 1]

            if column_name in [
                "Monthly Fixed Fee (from the Price Book)",
                "Billed Amount (Rate)",
                "Billed Amount Numeric",
            ]:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00;-'

            if column_name == "Difference (Billed Amount - Monthly Fixed Fee)":
                if monthly_fee_col_letter and billed_rate_col_letter:
                    cell.value = (
                        f'=IF(OR({monthly_fee_col_letter}{row_index}="",'
                        f'{billed_rate_col_letter}{row_index}=""),"",'
                        f'{billed_rate_col_letter}{row_index}-{monthly_fee_col_letter}{row_index})'
                    )
                cell.number_format = '$#,##0.00;$#,##0.00;-'

        # Status highlighting, if this sheet has Match Status
        if status_col_index:
            status_cell = ws.cell(row=row_index, column=status_col_index)
            status = status_cell.value

            if status == "Matched":
              status_cell.fill = matched_fill

            elif status == "Missing from Price Book":
              status_cell.fill = missing_fill

            elif status == "Missing Service ID from Report":
               status_cell.fill = missing_service_fill

    # Column widths
    for col_index, header in enumerate(sheet_columns, start=1):
        width = max(14, min(42, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(col_index)].width = width

    # Override widths for comparison sheets
    if sheet_name != "Missing Service ID from Report":
        widths = {
            1: 26,
            2: 30,
            3: 24,
            4: 32,
            5: 24,
            6: 24,
            7: 22,
            8: 18,
            9: 38,
        }

        for col_index, width in widths.items():
            if col_index <= len(sheet_columns):
                ws.column_dimensions[get_column_letter(col_index)].width = width

    ws.row_dimensions[1].height = 36

    # Add Excel table
    max_row = max(1, len(df) + 1)
    max_col = len(sheet_columns)
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

    if max_row >= 2 and max_col >= 1:
        base_table_name = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name[:20]).strip("_")

        if not base_table_name:
            base_table_name = f"Table{sheet_index + 1}"

        table_name = base_table_name
        suffix = 1
        while table_name in used_table_names:
            suffix += 1
            table_name = f"{base_table_name}_{suffix}"

        used_table_names.add(table_name)

        table = Table(displayName=table_name, ref=table_ref)

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style
        ws.add_table(table)

# Highlight non-zero differences only on sheets that have the Difference column
    if "Difference (Billed Amount - Monthly Fixed Fee)" in sheet_columns and max_row >= 2:

      difference_col = get_column_letter(
        sheet_columns.index(
            "Difference (Billed Amount - Monthly Fixed Fee)"
        ) + 1
      )

      red_fill = PatternFill(
        "solid",
        fgColor="FFC7CE"
      )

      red_font = Font(
        color="E06666",
        bold=True
      )

      ws.conditional_formatting.add(
        f"{difference_col}2:{difference_col}{max_row}",
        FormulaRule(
            formula=[
                f'AND(ISNUMBER({difference_col}2),ABS({difference_col}2)>0.005)'
            ],
            fill=red_fill,
            font=red_font,
        ),
      )

# =========================
# Add Summary sheet
# =========================

summary = wb.create_sheet("Summary")
summary.sheet_view.showGridLines = False

summary["A1"] = "Summary"
summary["A1"].font = Font(bold=True, size=14, color="1F4E78")

summary_rows = [
    (
        "Matched Service IDs",
        len(comparison[comparison["Match Status (Service ID)"] == "Matched"]),
    ),
    (
        "Missing from Price Book",
        len(comparison[comparison["Match Status (Service ID)"] == "Missing from Price Book"]),
    ),
    (
        "Missing Service ID from Report",
        len(missing_service_id_report),
    ),
    (
        "Mismatched Rate",
        len(mismatched_rate),
    ),
    (
        "Rows removed because billed amount is zero",
        zero_billed_count,
    ),
    ("Total comparison rows", len(comparison)),
]

for row_index, (label, value) in enumerate(summary_rows, start=3):
    summary.cell(row=row_index, column=1, value=label).font = Font(bold=True)
    summary.cell(row=row_index, column=2, value=value)

summary.column_dimensions["A"].width = 46
summary.column_dimensions["B"].width = 18


# =========================
# Save workbook
# =========================

wb.save(output_file)

print(f"Created: {output_file}")
print(f"Total rows: {len(comparison)}")
print("Matched:", len(comparison[comparison["Match Status (Service ID)"] == "Matched"]))
print("Missing from Price Book:", len(comparison[comparison["Match Status (Service ID)"] == "Missing from Price Book"]))
print("Missing Service ID:", len(missing_service_id_report))
print("Rows removed because billed amount is zero:", zero_billed_count)
