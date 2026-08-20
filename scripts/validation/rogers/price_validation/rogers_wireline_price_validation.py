# wireline price validation

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

DATA_DIR='.'

data_price_xlsx=os.path.join(DATA_DIR,'data.xlsx')
voice_price_xlsx=os.path.join(DATA_DIR,'voice.xlsx')
report_xlsx=os.path.join(DATA_DIR,'BC_GOV_Wireline_report_20260531_Consolidated.xlsx')
output_file=os.path.join(DATA_DIR,'Wireline_Service_ID_Comparison_Rate.xlsx')



def find_col(columns, options):
    normalized={re.sub(r'[^a-z0-9]','',str(c).lower()):c for c in columns}
    for option in options:
        key=re.sub(r'[^a-z0-9]','',option.lower())
        if key in normalized:
            return normalized[key]
    for c in columns:
        if any(o.lower() in str(c).lower() for o in options):
            return c
    raise KeyError(f'Missing column among: {options}')


def parse_money(value):
    if pd.isna(value):
        return None
    text=str(value).replace(',','').strip()
    m=re.search(r'([-+]?[0-9]+(?:\.[0-9]+)?)', text)
    return float(m.group(1)) if m else None

# Load files
price_data=pd.read_excel(data_price_xlsx, engine='openpyxl')
price_voice=pd.read_excel(voice_price_xlsx, engine='openpyxl')
price=pd.concat([price_data,price_voice], ignore_index=True)
report=pd.read_excel(report_xlsx, dtype=str, engine='openpyxl')

price.columns=[str(c).strip() for c in price.columns]
report.columns=[str(c).strip() for c in report.columns]

# Columns
price_service_id_col=find_col(price.columns,['Service ID'])
price_fee_col=find_col(price.columns,['Monthly Fixed Fee'])
report_service_id_col=find_col(report.columns,['SERVICE_ID','Service ID'])
report_rate_col=find_col(report.columns,['RATE'])
report_billed_col=find_col(report.columns,['BILLED_AMOUNT(PRE-TAX)'])

report_productline_col=find_col(report.columns,['PRODUCTLINE'])
report_quantity_col = find_col(
    report.columns,
    ["QUANTITY"]
)



# Price book prep
price['Normalized Service ID']=price[price_service_id_col].astype(str).str.strip()
price['Monthly Fixed Fee Numeric']=price[price_fee_col].map(parse_money)
price_clean = price[
    [price_service_id_col,
     'Normalized Service ID',
     'Monthly Fixed Fee Numeric'
    ]
].copy()
# Report prep
report['Normalized Service ID']=report[report_service_id_col].astype(str).str.strip()
report['Report Rate']=pd.to_numeric(report[report_rate_col].astype(str).str.replace(',',''), errors='coerce')
def extract_amount(value):
    if pd.isna(value):
        return None

    text = str(value).strip()

    if text == "" or text.lower() in ["nan", "none", "null"]:
        return None

    text = text.replace(",", "")
    text = text.replace("$", "")

    # Handle accounting negatives like (2.50)
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    return pd.to_numeric(text, errors="coerce")

# Find columns for charge description and type
report_charge_description_col = find_col(
    report.columns,
    ["CHARGE_DESCRIPTION"]
)

report_charge_type_col = find_col(
    report.columns,
    ["CHARGETYPE"]
)
# Create clean numeric billed amount column
report["Billed Amount Numeric"] = (
    report[report_billed_col]
    .apply(extract_amount)
    .astype(float)
)

print(
    "\nRows before filtering:",
    len(report)
)

print(
    "Rows with non-zero billed amount:",
    len(
        report[
            report["Billed Amount Numeric"]
            .fillna(0)
            .abs() > 0.005
        ]
    )
)

# Remove zero billed amount rows
report = report[
    report["Billed Amount Numeric"]
    .fillna(0)
    .abs() > 0.005
].copy()

# Capture rows with blank Service IDs and non-zero billed amount
missing_service_id = report[
    report[report_service_id_col]
        .fillna("")
        .astype(str)
        .str.strip()
        .eq("")
].copy()

missing_service_id = missing_service_id[
    missing_service_id["Billed Amount Numeric"]
        .fillna(0)
        .abs() > 0.005
].copy()


# Remove rows with blank Service IDs
report = report[
    report[report_service_id_col]
    .fillna("")
    .astype(str)
    .str.strip()
    .ne("")
].copy()


monthly_report_clean = report.copy()

# Remove the original report billed column to avoid duplicate/conflicting columns
monthly_report_clean.drop(
    columns=[
        report_billed_col,
        "Billed Amount (Pre-Tax)"
    ],
    inplace=True,
    errors="ignore"
)

# Use only the clean numeric billed amount in the output
monthly_report_clean["Billed Amount (Pre-Tax)"] = (
    monthly_report_clean["Billed Amount Numeric"]
)

if "Report Row" not in monthly_report_clean.columns:
    monthly_report_clean.insert(
        0,
        "Report Row",
        range(1, len(monthly_report_clean) + 1)
    )

monthly_report_clean.rename(
    columns={
        report_service_id_col: "Monthly Report Service ID",
        report_productline_col: "Product Line",
        report_quantity_col: "Quantity"
    },
    inplace=True
)




comparison = monthly_report_clean.merge(
    price_clean[
        [
            price_service_id_col,
            "Normalized Service ID",
            "Monthly Fixed Fee Numeric",
        ]
    ],
    on="Normalized Service ID",
    how="left"
)
print("Comparison rows:", len(comparison))






comparison.rename(columns={
 price_service_id_col:'Price Book Service ID',
 'Monthly Fixed Fee Numeric':'Monthly Fixed Fee (from the Price Book)'
}, inplace=True)

missing_service_id.rename(
    columns={
        report_service_id_col: "Monthly Report Service ID",
        report_productline_col: "Product Line",
        report_quantity_col: "Quantity",
        report_charge_type_col: "Charge Type",
        report_charge_description_col: "Charge Description"
    },
    inplace=True
)
missing_service_id["Match Status (Service ID)"] = "Missing Service ID from Report"

# No Service ID means no match in the price book
missing_service_id["Price Book Service ID"] = None
missing_service_id["Monthly Fixed Fee (from the Price Book)"] = None

# Keep report values
missing_service_id["Report Rate"] = pd.to_numeric(
    missing_service_id[report_rate_col].astype(str).str.replace(",", ""),
    errors="coerce"
)

missing_service_id["Billed Amount (Pre-Tax)"] = (
    missing_service_id["Billed Amount Numeric"]
)

# No difference can be calculated
missing_service_id["Difference (Rate - Monthly Fixed Fee)"] = None


def status(row):
    p=row['Monthly Fixed Fee (from the Price Book)']
    r=row['Report Rate']
    if pd.isna(p):
        return 'Missing from Price Book'
    if pd.isna(r):
        return 'Missing Report Rate'
    if abs(r-p)<=0.005:
        return 'Matched'
    return 'Rate Mismatch'

comparison['Match Status (Service ID)']=comparison.apply(status, axis=1)
comparison['Difference (Rate - Monthly Fixed Fee)']=None


columns = [
    'Price Book Service ID',
    'Monthly Report Service ID',
    'Product Line',
    'Match Status (Service ID)',
    'Monthly Fixed Fee (from the Price Book)',
    'Report Rate',
    'Quantity',
    'Billed Amount (Pre-Tax)',
    'Difference (Rate - Monthly Fixed Fee)'
]


# print("\nMissing Service ID columns:")
# print(missing_service_id.columns.tolist())
missing_service_id_for_comparison = (
    missing_service_id[columns].copy()
)

comparison=comparison[columns]

comparison = pd.concat(
    [
        comparison,
        missing_service_id_for_comparison
    ],
    ignore_index=True
)



wb=Workbook()

views=[
 ('Complete comparison',comparison),

 ('Matched',
  comparison[
      comparison['Match Status (Service ID)']=='Matched'
  ]),

 ('Rate Mismatch',
  comparison[
      comparison['Match Status (Service ID)']=='Rate Mismatch'
  ]),

 ('Missing from Price Book',
  comparison[
      comparison['Match Status (Service ID)']=='Missing from Price Book'
  ])
]

header_fill=PatternFill('solid', fgColor='1F4E78')
header_font=Font(color='FFFFFF', bold=True)

for idx,(name,df) in enumerate(views):
    ws=wb.active if idx==0 else wb.create_sheet(name[:31])
    ws.title=name[:31]
    for c,h in enumerate(columns,1):
        cell=ws.cell(1,c,h)
        cell.fill=header_fill
        cell.font=header_font
    for r,row in enumerate(df.itertuples(index=False),2):
        for c,val in enumerate(row,1):
            ws.cell(r,c,val)
        ws.cell(r,9,f'=IF(OR(E{r}="",F{r}=""),"",F{r}-E{r})')


# Missing Service ID sheet
ws = wb.create_sheet("Missing Service ID from Report")
missing_columns = list(missing_service_id.columns)

# Header row
for c, h in enumerate(missing_columns, 1):
    cell = ws.cell(1, c, h)
    cell.fill = header_fill
    cell.font = header_font

# Data rows
for r, row in enumerate(
        missing_service_id[missing_columns]
        .itertuples(index=False),
        start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(r, c, val)

for c, h in enumerate(missing_columns, 1):
    cell = ws.cell(1, c, h)
    cell.fill = header_fill
    cell.font = header_font

for r, row in enumerate(
        missing_service_id[missing_columns]
        .itertuples(index=False),
        start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(r, c, val)

summary=wb.create_sheet('Summary')
summary['A1']='Summary'
summary['A3']='Matched'; summary['B3']=len(comparison[comparison['Match Status (Service ID)']=='Matched'])
summary['A4']='Rate Mismatch'; summary['B4']=len(comparison[comparison['Match Status (Service ID)']=='Rate Mismatch'])
summary['A5']='Missing from Price Book'; summary['B5']=len(comparison[comparison['Match Status (Service ID)']=='Missing from Price Book'])
summary['A6'] = 'Missing Service ID from Report'
summary['B6'] = len(
    comparison[
        comparison['Match Status (Service ID)']
        == 'Missing Service ID from Report'
    ]
)

summary['A7']='Total Rows'
summary['B7']=len(comparison)





wb.save(output_file)
print('Created', output_file)

