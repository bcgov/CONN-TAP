"""
All BGE Dashboard Report Generator
==================================

Business rules
--------------
- Voice, IVR, Hosted IVR, Long Distance, and Conferencing are rolled into Voice.
- Out of Scope is excluded from consolidated reporting.
- Other is excluded from consolidated reporting.
- School Districts is normalized to SD.
- TSMA and TSMA Lite are treated as Telus-side spend.
- TELUS NGTA is treated as Telus-side spend.
- Rogers NGTA is treated as Rogers-side spend.

How to use
----------
1. Put this script in the same folder as your latest dashboard workbook.
2. If needed, update SOURCE_FILE below to match the workbook filename.
3. Run:
      python bge_dashboard_report_clean.py
4. The output workbook is created as:
      All_BGE_Dashboard_Report.xlsx
"""

from pathlib import Path
from collections import defaultdict
import re

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# =====================================================================
# CONFIGURATION
# =====================================================================

SOURCE_FILE = "spend_tracking_2026_07_03_0826AM.xlsx"
OUTPUT_FILE = "All_BGE_Report.xlsx"

SHEET_SOURCE = "Sheet1"
SHEET_BGE_LIST = "BGE_List"
SHEET_DETAIL_DATA = "BGE_Category_Detail_Data"
SHEET_CATEGORY_DASHBOARD = "BGE_Category_Dashboard"
SHEET_PROVIDER_DASHBOARD = "BGE_Spend_by_Provider_Dashboard"
SHEET_GOVBC_DASHBOARD = "GovBC_Dashboard"
SHEET_README = "README_BGE_Dashboard"

OLD_SHEET_NAMES = [
    "BGE_Dashboard",
    "GovBC_Stacked_bar_charts",
]

CUTOFF_MONTH = "Mar 2026"

TELUS_GREEN = "00B050"
ROGERS_RED = "FF0000"
TOTAL_BLUE = "4472C4"
DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "5B9BD5"
LIGHT_BLUE = "BDD7EE"
DATA_ORANGE = "ED7D31"
VOICE_GREEN = "70AD47"
DARK_GREY = "404040"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
LIGHT_FILL = PatternFill("solid", fgColor="F7FBFF")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
WHITE_FONT = Font(color="FFFFFF")
THIN_BLUE = Side(style="thin", color="B7CEE8")

CURFMT = "$#,##0;[Red]($#,##0);-"
CHART_AXIS_CURFMT = '$#,##0;-$#,##0'

EXPECTED_BGES = [
    "BC Hydro", "BCLC", "ECC", "FHA", "FNHA", "Gov BC", "ICBC", "IHA",
    "NHA", "PHSA", "SD", "VCHA", "VIHA", "WSBC"
]

DETAIL_FIELDS = [
    "Cellular_TSMA", "Data_TSMA", "Voice_TSMA",
    "Plan_Telus", "Plan_Rogers",
    "HW_Telus", "HW_Rogers",
    "Data_Telus", "Data_Rogers",
    "Voice_Telus", "Voice_Rogers",
]

TELUS_FIELDS = [
    "Cellular_TSMA",
    "Data_TSMA",
    "Voice_TSMA",
    "Plan_Telus",
    "HW_Telus",
    "Data_Telus",
    "Voice_Telus",
]

ROGERS_FIELDS = [
    "Plan_Rogers",
    "HW_Rogers",
    "Data_Rogers",
    "Voice_Rogers",
]


# =====================================================================
# BASIC HELPERS
# =====================================================================

def to_number(value):
    """Convert workbook values to float, treating blanks and Excel errors as 0."""
    try:
        if value is None or value == "":
            return 0.0
        if isinstance(value, str) and value.startswith("#"):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def norm_bge(name):
    """Normalize BGE names for dashboard use."""
    return "SD" if name == "School Districts" else name


def norm_text(value):
    """Normalize text for comparisons."""
    return str(value or "").strip().lower()


def find_source_file():
    """Find the source workbook in the current directory."""
    src = Path(SOURCE_FILE)
    if src.exists():
        return src

    files = sorted(Path(".").glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for file_path in files:
        if "dashboard" in file_path.name.lower():
            return file_path

    raise FileNotFoundError(
        "Could not find the dashboard workbook. Put this script beside the workbook or update SOURCE_FILE."
    )


def get_cutoff_month_index(months, cutoff_month=CUTOFF_MONTH):
    """Return cutoff month index, or the last available month if cutoff is not found."""
    try:
        return months.index(cutoff_month)
    except ValueError:
        return len(months) - 1


def get_last_month_with_data(months, agg):
    """Return the last month index that contains any non-zero spend."""
    last_idx = 0

    for idx in range(len(months)):
        total = sum(abs(values[idx]) for values in agg.values())
        if total > 0:
            last_idx = idx

    return last_idx


def move_sheet_to_front(wb, sheet_name):
    """
    Move a worksheet to the first tab position.

    Note: openpyxl does not expose a public move-to-front API, so this uses the
    workbook internal sheet list. This is commonly used for controlled reports.
    """
    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    wb._sheets.insert(0, ws)


# =====================================================================
# STYLE HELPERS
# =====================================================================

def style_header_cell(cell, wrap_text=True):
    """Apply standard dashboard header style."""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", wrap_text=wrap_text)


def style_title_cell(cell, size=14):
    """Apply standard dashboard title style."""
    cell.font = Font(bold=True, size=size, color="1F4E78")


def style_data_row(ws, row, start_col, end_col):
    """Apply alternating fill and bottom border to a data row."""
    if row % 2 == 0:
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).fill = LIGHT_FILL

    for col in range(start_col, end_col + 1):
        ws.cell(row, col).border = Border(bottom=THIN_BLUE)


def set_column_widths(ws, widths):
    """Apply worksheet column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def hide_helper_columns(ws, start_col, end_col):
    """Make helper columns practically invisible but still usable for charts."""
    for col_idx in range(start_col, end_col + 1):
        col = get_column_letter(col_idx)
        ws.column_dimensions[col].width = 0.1
        for cell in ws[col]:
            cell.font = WHITE_FONT
            cell.fill = WHITE_FILL


# =====================================================================
# CHART HELPERS
# =====================================================================

def format_currency_axis(chart_obj):
    """Force chart Y-axis to use a consistent currency format."""
    try:
        chart_obj.y_axis.numFmt = CHART_AXIS_CURFMT
    except Exception:
        pass


def set_chart_base_style(chart_obj, height=16, width=32, legend_position="t"):
    """Apply common chart layout settings."""
    chart_obj.height = height
    chart_obj.width = width
    chart_obj.x_axis.delete = False
    chart_obj.y_axis.delete = False

    try:
        chart_obj.legend.position = legend_position
    except Exception:
        pass

    format_currency_axis(chart_obj)


def apply_line_colors(chart_obj, colors):
    """Apply line colors to chart series."""
    for series, color in zip(chart_obj.series, colors):
        series.graphicalProperties.line.solidFill = color


def apply_bar_colors(chart_obj, colors, border_color=None):
    """Apply fill and line colors to bar chart series."""
    for series, color in zip(chart_obj.series, colors):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = border_color or color
        try:
            series.invertIfNegative = False
            series.graphicalProperties.noFill = False
        except Exception:
            pass


def create_reference(ws, min_col, max_col, min_row, max_row):
    """Create an openpyxl Reference with readable parameter names."""
    return Reference(
        ws,
        min_col=min_col,
        max_col=max_col,
        min_row=min_row,
        max_row=max_row,
    )


# =====================================================================
# SOURCE PARSING
# =====================================================================

def get_month_cols(ws):
    """Read all month columns from Sheet1 row 4."""
    months, cols = [], []
    pattern = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec) \d{4}$")

    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(4, col_idx).value
        if isinstance(value, str) and pattern.match(value.strip()):
            month = value.strip().replace("Sept", "Sep")
            months.append(month)
            cols.append(col_idx)

    return months, cols


def classify(label, source):
    """Classify a source row label into dashboard category fields."""
    text = norm_text(label)
    if not text:
        return None

    if text.startswith("total") or "out of scope" in text or text == "other" or text.startswith("other"):
        return None

    if source in ("TSMA", "TSMA Lite"):
        if "cellular" in text or "mms" in text:
            return "Cellular_TSMA"
        if "data" in text:
            return "Data_TSMA"
        if "voice" in text or "ivr" in text or "long distance" in text or "conferencing" in text:
            return "Voice_TSMA"
        return None

    if "cellular h/w" in text or "cellular hw" in text or "h/w" in text:
        return "Cellular_HW"
    if "cellular plan" in text or "cellular plans" in text or text == "cellular" or "mms" in text:
        return "Cellular_Plan"
    if "data" in text:
        return "Data"
    if "voice" in text or "ivr" in text or "long distance" in text or "conferencing" in text:
        return "Voice"

    return None


def find_row_col_b(ws, text, start=1):
    """Find the first row where column B equals text."""
    for row_idx in range(start, ws.max_row + 1):
        if str(ws.cell(row_idx, 2).value).strip() == text:
            return row_idx
    return None


def parse_rows(ws, start, end, source, month_cols):
    """Parse rows within a TSMA, TELUS, or Rogers block."""
    rows = []
    current_bge = None

    for row_idx in range(start, end):
        aval = ws.cell(row_idx, 1).value
        bval = ws.cell(row_idx, 2).value

        if aval is not None and str(aval).strip() not in ("", "BGEs"):
            aval_clean = str(aval).strip()
            if not (aval_clean.startswith("(+") and current_bge):
                current_bge = aval_clean

        if not current_bge:
            continue

        category = classify(bval, source)
        if category is None:
            continue

        values = [to_number(ws.cell(row_idx, col_idx).value) for col_idx in month_cols]
        rows.append((norm_bge(current_bge), source, category, values))

    return rows


def parse_tsma(ws, month_cols):
    """Parse the first TSMA block."""
    bge_row = None

    for row_idx in range(1, ws.max_row + 1):
        if str(ws.cell(row_idx, 1).value).strip() == "BGEs":
            bge_row = row_idx
            break

    if not bge_row:
        return []

    end = ws.max_row + 1
    for row_idx in range(bge_row + 1, ws.max_row + 1):
        value = ws.cell(row_idx, 2).value
        if isinstance(value, str) and value.startswith("TOTAL "):
            end = row_idx
            break

    return parse_rows(ws, bge_row + 1, end, "TSMA", month_cols)


def parse_provider(ws, provider, month_cols):
    """Parse a provider block, e.g. TELUS NGTA or Rogers NGTA."""
    start = find_row_col_b(ws, provider)
    if start is None:
        return []

    bge_row = None
    for row_idx in range(start, ws.max_row + 1):
        if str(ws.cell(row_idx, 1).value).strip() == "BGEs":
            bge_row = row_idx
            break

    if bge_row is None:
        return []

    end = ws.max_row + 1
    for row_idx in range(bge_row + 1, ws.max_row + 1):
        label = str(ws.cell(row_idx, 2).value or "").strip()
        if label.startswith("TOTAL ") or label in ("Rogers NGTA", "Out of Scope", "TSMA Lite"):
            end = row_idx
            break

    return parse_rows(ws, bge_row + 1, end, provider, month_cols)


def map_category_to_field(source, category):
    """Map parsed source/category values to detail field names."""
    if category in ("Cellular_TSMA", "Data_TSMA", "Voice_TSMA"):
        return category

    mapping = {
        ("TELUS NGTA", "Cellular_Plan"): "Plan_Telus",
        ("Rogers NGTA", "Cellular_Plan"): "Plan_Rogers",
        ("TELUS NGTA", "Cellular_HW"): "HW_Telus",
        ("Rogers NGTA", "Cellular_HW"): "HW_Rogers",
        ("TELUS NGTA", "Data"): "Data_Telus",
        ("Rogers NGTA", "Data"): "Data_Rogers",
        ("TELUS NGTA", "Voice"): "Voice_Telus",
        ("Rogers NGTA", "Voice"): "Voice_Rogers",
    }
    return mapping.get((source, category))


def build_detail_data(src_ws):
    """Build detailed BGE/category data from Sheet1."""
    months, month_cols = get_month_cols(src_ws)
    agg = defaultdict(lambda: [0.0] * len(months))

    all_rows = []
    all_rows.extend(parse_tsma(src_ws, month_cols))
    all_rows.extend(parse_provider(src_ws, "TELUS NGTA", month_cols))
    all_rows.extend(parse_provider(src_ws, "Rogers NGTA", month_cols))

    for bge, source, category, values in all_rows:
        field = map_category_to_field(source, category)
        if field:
            agg[(bge, field)] = [a + b for a, b in zip(agg[(bge, field)], values)]

    return months, agg, DETAIL_FIELDS


# =====================================================================
# BGE LIST AND DROPDOWNS
# =====================================================================

def get_bges(wb, agg):
    """Get BGE list from BGE_List, falling back to parsed data."""
    if SHEET_BGE_LIST in wb.sheetnames:
        ws = wb[SHEET_BGE_LIST]
        values = [ws.cell(row_idx, 1).value for row_idx in range(2, ws.max_row + 1)]
        values = [norm_bge(str(value).strip()) for value in values if value]

        if values:
            seen = set(values)
            ordered = [bge for bge in EXPECTED_BGES if bge in seen]
            extras = [bge for bge in values if bge not in EXPECTED_BGES and bge != "All BGEs"]
            final_bges = ordered + extras

            if "All BGEs" not in final_bges:
                final_bges.insert(0, "All BGEs")

            return final_bges

    parsed_bges = sorted({key[0] for key in agg.keys()})
    ordered = [bge for bge in EXPECTED_BGES if bge in parsed_bges]
    extras = [bge for bge in parsed_bges if bge not in ordered]

    return ["All BGEs"] + ordered + extras


def rebuild_bge_list(wb, bges):
    """Rebuild hidden BGE_List sheet with complete BGE list."""
    if SHEET_BGE_LIST in wb.sheetnames:
        ws = wb[SHEET_BGE_LIST]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SHEET_BGE_LIST)

    ws.append(["BGE"])
    for bge in bges:
        ws.append([bge])

    style_header_cell(ws["A1"], wrap_text=False)
    ws.column_dimensions["A"].width = 22
    ws.sheet_state = "hidden"
    return ws


def add_bge_dropdown(ws, cell_address, bges, include_all_bges=True):
    """Add a BGE dropdown using the hidden BGE_List sheet."""
    if include_all_bges:
        first_row = 2
    else:
        first_row = 3 if bges and bges[0] == "All BGEs" else 2

    formula = f"={quote_sheetname(SHEET_BGE_LIST)}!$A${first_row}:$A${len(bges) + 1}"

    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.error = "Select a BGE from the list."
    dv.errorTitle = "Invalid BGE"
    dv.prompt = "Choose a BGE from the dropdown."
    dv.promptTitle = "Select BGE"

    ws.add_data_validation(dv)
    dv.add(ws[cell_address])


def style_dropdown_cell(cell):
    """Apply standard formatting to a dashboard dropdown cell."""
    cell.fill = SUB_FILL
    cell.font = Font(bold=True, color="1F4E78")
    cell.alignment = Alignment(horizontal="center")


def repair_dashboard_dropdowns(wb, bges):
    """Repair BGE dropdowns in dashboard sheets."""
    if SHEET_BGE_LIST not in wb.sheetnames:
        rebuild_bge_list(wb, bges)

    dashboard_sheets = [
        SHEET_CATEGORY_DASHBOARD,
        SHEET_PROVIDER_DASHBOARD,
    ]

    for sheet_name in dashboard_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        try:
            ws.data_validations.dataValidation = []
        except Exception:
            pass

        add_bge_dropdown(ws, "B2", bges, include_all_bges=True)

        current = ws["B2"].value
        if current not in bges:
            ws["B2"] = "Gov BC" if "Gov BC" in bges else bges[0]

        style_dropdown_cell(ws["B2"])


# =====================================================================
# DETAIL DATA SHEET
# =====================================================================

def ensure_data_sheet(wb, months, bges, agg, fields):
    """Create hidden detail data sheet used by dashboard formulas."""
    if SHEET_DETAIL_DATA in wb.sheetnames:
        del wb[SHEET_DETAIL_DATA]

    ws = wb.create_sheet(SHEET_DETAIL_DATA)
    headers = ["BGE", "MonthIndex", "Month"] + fields + ["Key"]
    ws.append(headers)

    month_count = len(months)

    for bge in bges:
        if bge == "All BGEs":
            continue

        for month_index, month in enumerate(months, start=1):
            row = [bge, month_index, month]
            for field in fields:
                row.append(agg.get((bge, field), [0.0] * month_count)[month_index - 1])
            row.append(f"{bge}|{month_index}")
            ws.append(row)

    for month_index, month in enumerate(months, start=1):
        row = ["All BGEs", month_index, month]

        for field in fields:
            total = 0.0
            for bge in bges:
                if bge == "All BGEs":
                    continue
                total += agg.get((bge, field), [0.0] * month_count)[month_index - 1]
            row.append(total)

        row.append(f"All BGEs|{month_index}")
        ws.append(row)

    for cell in ws[1]:
        style_header_cell(cell)

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    try:
        table = Table(
            displayName="BGECategoryDetailDataTable",
            ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    except Exception:
        pass

    ws.sheet_state = "hidden"
    return ws


# =====================================================================
# FORMULA HELPERS
# =====================================================================

def detail_match_formula(bge_expression, month_index):
    """Return a MATCH formula expression for BGE_Category_Detail_Data Key column."""
    return f'MATCH({bge_expression}&"|"&{month_index},{SHEET_DETAIL_DATA}!$O:$O,0)'


def detail_index_formula(column_letter, match_expression):
    """Return an INDEX formula expression for BGE_Category_Detail_Data."""
    return f'INDEX({SHEET_DETAIL_DATA}!${column_letter}:${column_letter},{match_expression})'


def detail_value_expr(column_letter, match_expression):
    """Return an IFERROR-wrapped detail value expression."""
    return f'IFERROR({detail_index_formula(column_letter, match_expression)},0)'


def sum_detail_expr(column_letters, match_expression):
    """Return formula expression that sums several detail data columns."""
    return "+".join(detail_value_expr(col, match_expression) for col in column_letters)


def zero_as_dash_formula(expression):
    """Return Excel formula that shows dash for zero values."""
    return f'=IF(({expression})=0,"-",({expression}))'


# =====================================================================
# ROLLUP HELPERS
# =====================================================================

def get_month_value(agg, bge, field, month_idx, month_count):
    """Safely get a monthly value from the aggregation dictionary."""
    return agg.get((bge, field), [0.0] * month_count)[month_idx]


def build_all_bge_rollup(months, agg, bges):
    """
    Build monthly All-BGE rollup.

    Telus  = TSMA + TSMA Lite + TELUS NGTA
    Rogers = Rogers NGTA
    Total  = Telus + Rogers
    """
    rollup = []
    month_count = len(months)

    for month_idx, month in enumerate(months):
        telus = 0.0
        rogers = 0.0

        for bge in bges:
            if bge == "All BGEs":
                continue

            telus += sum(get_month_value(agg, bge, field, month_idx, month_count) for field in TELUS_FIELDS)
            rogers += sum(get_month_value(agg, bge, field, month_idx, month_count) for field in ROGERS_FIELDS)

        rollup.append({"Month": month, "Telus": telus, "Rogers": rogers, "Total": telus + rogers})

    return rollup


# =====================================================================
# BGE CATEGORY DASHBOARD
# =====================================================================

def rebuild_category_dashboard(wb, months, bges, last_month_idx):
    """Rebuild BGE_Category_Dashboard with detailed table and stacked category chart."""
    if SHEET_CATEGORY_DASHBOARD in wb.sheetnames:
        idx = wb.sheetnames.index(SHEET_CATEGORY_DASHBOARD)
        del wb[SHEET_CATEGORY_DASHBOARD]
        ws = wb.create_sheet(SHEET_CATEGORY_DASHBOARD, idx)
    else:
        ws = wb.create_sheet(SHEET_CATEGORY_DASHBOARD, 1)

    ws.sheet_view.showGridLines = False
    march_cutoff_idx = get_cutoff_month_index(months)

    ws["A1"] = "BGE Monthly Category Dashboard"
    style_title_cell(ws["A1"], size=16)
    ws["A2"] = "Select BGE:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = "Gov BC" if "Gov BC" in bges else bges[0]
    style_dropdown_cell(ws["B2"])
    ws["D2"] = "Detailed monthly table and stacked category bars update from the selected BGE."
    ws["D2"].font = Font(italic=True, color="666666")
    ws["F44"] = '=IF($B$2="","",$B$2&" Monthly Spend by Service Tower")'
    style_title_cell(ws["F44"], size=12)

    add_bge_dropdown(ws, "B2", bges, include_all_bges=True)

    headers = [
        "Month",
        "Cellular (TSMA+TSMA Lite)",
        "Data (TSMA+TSMA Lite)",
        "Voice (TSMA+TSMA Lite)",
        "Cellular Plan (TELUS NGTA)",
        "Cellular Plan (Rogers NGTA)",
        "Total Cellular Plan",
        "Cellular H/W (TELUS NGTA)",
        "Cellular H/W (Rogers NGTA)",
        "Total Cellular H/W",
        "Data (TELUS NGTA)",
        "Data (Rogers NGTA)",
        "Total Data",
        "Voice (TELUS NGTA)",
        "Voice (Rogers NGTA)",
        "Total Voice",
    ]

    header_row = 4
    first_row = 5

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx)
        cell.value = header
        style_header_cell(cell)

    for month_index, _month in enumerate(months, start=1):
        row_idx = first_row + month_index - 1
        match = detail_match_formula("$B$2", month_index)

        formulas = [
            f'={detail_index_formula("C", match)}',
            zero_as_dash_formula(detail_value_expr("D", match)),
            zero_as_dash_formula(detail_value_expr("E", match)),
            zero_as_dash_formula(detail_value_expr("F", match)),
            zero_as_dash_formula(detail_value_expr("G", match)),
            zero_as_dash_formula(detail_value_expr("H", match)),
            zero_as_dash_formula(sum_detail_expr(["G", "H"], match)),
            zero_as_dash_formula(detail_value_expr("I", match)),
            zero_as_dash_formula(detail_value_expr("J", match)),
            zero_as_dash_formula(sum_detail_expr(["I", "J"], match)),
            zero_as_dash_formula(detail_value_expr("K", match)),
            zero_as_dash_formula(detail_value_expr("L", match)),
            zero_as_dash_formula(sum_detail_expr(["E", "K", "L"], match)),
            zero_as_dash_formula(detail_value_expr("M", match)),
            zero_as_dash_formula(detail_value_expr("N", match)),
            zero_as_dash_formula(sum_detail_expr(["F", "M", "N"], match)),
        ]

        for col_idx, formula in enumerate(formulas, start=1):
            ws.cell(row_idx, col_idx).value = formula
            if col_idx > 1:
                ws.cell(row_idx, col_idx).number_format = CURFMT

        style_data_row(ws, row_idx, 1, len(headers))

    widths = [14, 22, 22, 22, 22, 22, 18, 22, 22, 18, 18, 18, 16, 18, 18, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    helper_col = 18
    helper_headers = ["Month", "TSMA Cellular", "NGTA Cellular Plan", "NGTA Cellular H/W", "Data", "Voice"]

    for col_idx, header in enumerate(helper_headers, start=helper_col):
        ws.cell(header_row, col_idx).value = header

    for month_index, _month in enumerate(months, start=1):
        source_row = first_row + month_index - 1
        ws.cell(source_row, helper_col).value = f"=A{source_row}"
        ws.cell(source_row, helper_col + 1).value = f"=IFERROR(B{source_row},0)"
        ws.cell(source_row, helper_col + 2).value = f"=IFERROR(G{source_row},0)"
        ws.cell(source_row, helper_col + 3).value = f"=IFERROR(J{source_row},0)"
        ws.cell(source_row, helper_col + 4).value = f"=IFERROR(M{source_row},0)"
        ws.cell(source_row, helper_col + 5).value = f"=IFERROR(P{source_row},0)"

    helper_last_row = first_row + march_cutoff_idx

    category_chart = BarChart()
    category_chart.type = "col"
    category_chart.grouping = "stacked"
    category_chart.overlap = 100
    category_chart.gapWidth = 40
    category_chart.title = ""
    set_chart_base_style(category_chart, height=17, width=34, legend_position="t")

    data = create_reference(ws, helper_col + 1, helper_col + 5, header_row, helper_last_row)
    cats = Reference(ws, min_col=helper_col, min_row=first_row, max_row=helper_last_row)
    category_chart.add_data(data, titles_from_data=True)
    category_chart.set_categories(cats)
    apply_bar_colors(
        category_chart,
        [DARK_BLUE, MEDIUM_BLUE, LIGHT_BLUE, DATA_ORANGE, VOICE_GREEN],
        border_color=DARK_GREY,
    )

    hide_helper_columns(ws, helper_col, helper_col + 5)
    ws.add_chart(category_chart, "C45")

    ws.freeze_panes = "A3"
    return ws


# =====================================================================
# BGE SPEND BY PROVIDER DASHBOARD
# =====================================================================

def rebuild_bge_spend_by_provider_dashboard(wb, months, bges):
    """Rebuild BGE spend by provider dashboard."""
    if SHEET_PROVIDER_DASHBOARD in wb.sheetnames:
        idx = wb.sheetnames.index(SHEET_PROVIDER_DASHBOARD)
        del wb[SHEET_PROVIDER_DASHBOARD]
        ws = wb.create_sheet(SHEET_PROVIDER_DASHBOARD, idx)
    else:
        ws = wb.create_sheet(SHEET_PROVIDER_DASHBOARD, 0)

    ws.sheet_view.showGridLines = False

    ws["A1"] = "BGE Monthly Spend Dashboard"
    style_title_cell(ws["A1"], size=16)
    ws["A2"] = "Select BGE:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = "Gov BC" if "Gov BC" in bges else bges[0]
    style_dropdown_cell(ws["B2"])
    add_bge_dropdown(ws, "B2", bges, include_all_bges=True)

    ws["B6"] = '=IF($B$2="","",$B$2&" Spend by Provider")'
    style_title_cell(ws["B6"], size=12)
    ws["N6"] = '=IF($B$2="","",$B$2&" Spend by Provider")'
    style_title_cell(ws["N6"], size=12)

    # Section 1: selected BGE table
    ws["B4"] = '=IF($B$2="All BGEs","Selected BGE - Hidden for All BGEs","Selected BGE Monthly Spend")'
    style_title_cell(ws["B4"], size=14)
    ws["B5"] = '=IF($B$2="All BGEs","All BGEs selected - use the All BGEs Summary section below.","" )'
    ws["B5"].font = Font(bold=True, size=12, color="C00000")

    selected_start_row = 7
    selected_start_col = 2
    selected_headers = ["Month", "Total", "Telus", "Rogers"]

    for col_idx, header in enumerate(selected_headers, start=selected_start_col):
        cell = ws.cell(selected_start_row, col_idx)
        cell.value = header
        style_header_cell(cell, wrap_text=False)

    for month_index, _month in enumerate(months, start=1):
        row_idx = selected_start_row + month_index
        match = detail_match_formula("$B$2", month_index)
        telus_expr = sum_detail_expr(["D", "E", "F", "G", "I", "K", "M"], match)
        rogers_expr = sum_detail_expr(["H", "J", "L", "N"], match)
        total_expr = f"({telus_expr})+({rogers_expr})"

        ws.cell(row_idx, selected_start_col).value = (
            f'=IF($B$2="All BGEs","",{detail_index_formula("C", match)})'
        )
        ws.cell(row_idx, selected_start_col + 1).value = (
            f'=IF($B$2="All BGEs","",IF(({total_expr})=0,"-",({total_expr})))'
        )
        ws.cell(row_idx, selected_start_col + 2).value = (
            f'=IF($B$2="All BGEs","",IF(({telus_expr})=0,"-",({telus_expr})))'
        )
        ws.cell(row_idx, selected_start_col + 3).value = (
            f'=IF($B$2="All BGEs","",IF(({rogers_expr})=0,"-",({rogers_expr})))'
        )

        for col_idx in range(selected_start_col + 1, selected_start_col + 4):
            ws.cell(row_idx, col_idx).number_format = CURFMT

        style_data_row(ws, row_idx, selected_start_col, selected_start_col + 3)

    selected_last_row = selected_start_row + len(months)
    selected_chart_last_row = selected_last_row
    for idx, month in enumerate(months, start=1):
        if month == CUTOFF_MONTH:
            selected_chart_last_row = selected_start_row + idx
            break

    selected_chart = LineChart()
    selected_chart.title = ""
    set_chart_base_style(selected_chart, height=16, width=32, legend_position="t")
    selected_chart.add_data(
        create_reference(ws, selected_start_col + 1, selected_start_col + 3, selected_start_row, selected_chart_last_row),
        titles_from_data=True,
    )
    selected_chart.set_categories(
        Reference(ws, min_col=selected_start_col, min_row=selected_start_row + 1, max_row=selected_chart_last_row)
    )
    apply_line_colors(selected_chart, [TOTAL_BLUE, TELUS_GREEN, ROGERS_RED])
    ws.add_chart(selected_chart, "G7")

    # Section 2: All BGEs rollup table
    rollup_title_row = 57
    rollup_start_row = 58
    rollup_start_col = 2

    ws["N57"] = "All BGEs Spend by Provider"
    style_title_cell(ws["N57"], size=14)
    ws[f"B{rollup_title_row}"] = '=IF($B$2="All BGEs","All BGEs Summary (Selected)","All BGEs Combined")'
    style_title_cell(ws[f"B{rollup_title_row}"], size=14)
    ws["B41"] = '=IF($B$2="All BGEs","Showing consolidated results for ALL BGEs","" )'
    ws["B41"].font = Font(bold=True, size=12, color="C00000")
    ws["B43"] = '=IF($B$2="All BGEs","Use this section when All BGEs is selected","" )'
    ws["B43"].font = Font(bold=True, color="1F4E78")
    ws["B44"] = "Monthly Rollup Totals"
    ws["B44"].font = Font(bold=True)

    rollup_headers = ["Month", "Total", "Telus", "Rogers"]
    for col_idx, header in enumerate(rollup_headers, start=rollup_start_col):
        cell = ws.cell(rollup_start_row, col_idx)
        cell.value = header
        style_header_cell(cell, wrap_text=False)

    for month_index, _month in enumerate(months, start=1):
        row_idx = rollup_start_row + month_index
        match = detail_match_formula('"All BGEs"', month_index)
        telus_expr = sum_detail_expr(["D", "E", "F", "G", "I", "K", "M"], match)
        rogers_expr = sum_detail_expr(["H", "J", "L", "N"], match)
        total_expr = f"({telus_expr})+({rogers_expr})"

        ws.cell(row_idx, rollup_start_col).value = f'={detail_index_formula("C", match)}'
        ws.cell(row_idx, rollup_start_col + 1).value = zero_as_dash_formula(total_expr)
        ws.cell(row_idx, rollup_start_col + 2).value = zero_as_dash_formula(telus_expr)
        ws.cell(row_idx, rollup_start_col + 3).value = zero_as_dash_formula(rogers_expr)

        for col_idx in range(rollup_start_col + 1, rollup_start_col + 4):
            ws.cell(row_idx, col_idx).number_format = CURFMT

        style_data_row(ws, row_idx, rollup_start_col, rollup_start_col + 3)

    rollup_last_row = rollup_start_row + len(months)
    rollup_chart_last_row = rollup_last_row
    for idx, month in enumerate(months, start=1):
        if month == CUTOFF_MONTH:
            rollup_chart_last_row = rollup_start_row + idx
            break

    rollup_chart = LineChart()
    rollup_chart.title = ""
    set_chart_base_style(rollup_chart, height=16, width=32, legend_position="t")
    rollup_chart.add_data(
        create_reference(ws, rollup_start_col + 1, rollup_start_col + 3, rollup_start_row, rollup_chart_last_row),
        titles_from_data=True,
    )
    rollup_chart.set_categories(
        Reference(ws, min_col=rollup_start_col, min_row=rollup_start_row + 1, max_row=rollup_chart_last_row)
    )
    apply_line_colors(rollup_chart, [TOTAL_BLUE, TELUS_GREEN, ROGERS_RED])
    ws.add_chart(rollup_chart, "G58")

    set_column_widths(
        ws,
        {
            "A": 4,
            "B": 14,
            "C": 16,
            "D": 16,
            "E": 16,
            "L": 14,
            "M": 14,
            "N": 14,
            "O": 14,
        },
    )

    ws.freeze_panes = "A4"
    return ws


# =====================================================================
# GOV BC DASHBOARD
# =====================================================================

def rebuild_govbc_dashboard(wb, months):
    """Rebuild GovBC_Dashboard from scratch."""
    if SHEET_GOVBC_DASHBOARD in wb.sheetnames:
        idx = wb.sheetnames.index(SHEET_GOVBC_DASHBOARD)
        del wb[SHEET_GOVBC_DASHBOARD]
        ws = wb.create_sheet(SHEET_GOVBC_DASHBOARD, idx)
    else:
        ws = wb.create_sheet(SHEET_GOVBC_DASHBOARD)

    ws.sheet_view.showGridLines = False
    bge_name = "Gov BC"
    march_cutoff_idx = get_cutoff_month_index(months)

    ws["A1"] = "Gov BC Charts"
    style_title_cell(ws["A1"], size=16)
    ws["A2"] = "Source: Sheet1 -> BGE_Category_Detail_Data -> Gov BC only"
    ws["A2"].font = Font(italic=True, color="666666")

    # -----------------------------------------------------------------
    # Section 1: Cellular
    # -----------------------------------------------------------------
    cellular_header_row = 4
    cellular_first_row = cellular_header_row + 1

    ws["A3"] = "Gov BC Cellular Monthly Spend"
    style_title_cell(ws["A3"], size=14)
    ws["L3"] = "Gov BC Cellular Monthly Spend - Plan vs H/W"
    style_title_cell(ws["L3"], size=12)
    ws["L34"] = "Gov BC Cellular Monthly Spend by Provider"
    style_title_cell(ws["L34"], size=12)
    ws["L93"] = "Gov BC Voice Monthly Spend by Provider"
    style_title_cell(ws["L93"], size=12)
    ws["P138"] = "Gov BC Monthly Spend by Service Tower and Provider"
    style_title_cell(ws["P138"], size=12)

    cellular_headers = [
        "Month",
        "Cellular (TSMA+TSMA Lite)",
        "TELUS NGTA (Cellular Plan)",
        "TELUS NGTA (Cellular H/W)",
        "Rogers NGTA (Cellular Plan)",
        "Rogers NGTA (Cellular H/W)",
    ]

    for col_idx, header in enumerate(cellular_headers, start=1):
        cell = ws.cell(cellular_header_row, col_idx)
        cell.value = header
        style_header_cell(cell)

    for month_index, _month in enumerate(months, start=1):
        row_idx = cellular_first_row + month_index - 1
        match = detail_match_formula(f'"{bge_name}"', month_index)

        formulas = [
            f'={detail_index_formula("C", match)}',
            f'={detail_value_expr("D", match)}',
            f'={detail_value_expr("G", match)}',
            f'={detail_value_expr("I", match)}',
            f'={detail_value_expr("H", match)}',
            f'={detail_value_expr("J", match)}',
        ]

        for col_idx, formula in enumerate(formulas, start=1):
            ws.cell(row_idx, col_idx).value = formula
            if col_idx > 1:
                ws.cell(row_idx, col_idx).number_format = CURFMT

        style_data_row(ws, row_idx, 1, 6)

    cellular_last_row = cellular_first_row + march_cutoff_idx

    # Cellular helper table for stacked bar chart
    cellular_helper_col = 18
    cellular_helper_headers = ["Month", "Cellular Plan", "Cellular H/W"]

    for col_idx, header in enumerate(cellular_helper_headers, start=cellular_helper_col):
        ws.cell(cellular_header_row, col_idx).value = header

    for month_index, _month in enumerate(months, start=1):
        source_row = cellular_first_row + month_index - 1
        ws.cell(source_row, cellular_helper_col).value = f"=A{source_row}"
        ws.cell(source_row, cellular_helper_col + 1).value = (
            f"=IFERROR(B{source_row},0)+IFERROR(C{source_row},0)+IFERROR(E{source_row},0)"
        )
        ws.cell(source_row, cellular_helper_col + 2).value = (
            f"=IFERROR(D{source_row},0)+IFERROR(F{source_row},0)"
        )
        ws.cell(source_row, cellular_helper_col + 1).number_format = CURFMT
        ws.cell(source_row, cellular_helper_col + 2).number_format = CURFMT

    cellular_stack_chart = BarChart()
    cellular_stack_chart.type = "col"
    cellular_stack_chart.grouping = "stacked"
    cellular_stack_chart.overlap = 100
    cellular_stack_chart.gapWidth = 40
    cellular_stack_chart.title = ""
    set_chart_base_style(cellular_stack_chart, height=14, width=30, legend_position="t")
    cellular_stack_chart.add_data(
        create_reference(ws, cellular_helper_col + 1, cellular_helper_col + 2, cellular_header_row, cellular_last_row),
        titles_from_data=True,
    )
    cellular_stack_chart.set_categories(
        Reference(ws, min_col=cellular_helper_col, min_row=cellular_first_row, max_row=cellular_last_row)
    )
    apply_bar_colors(cellular_stack_chart, [TOTAL_BLUE, DATA_ORANGE])
    ws.add_chart(cellular_stack_chart, "I4")

    # Cellular line chart by provider
    trend_col = 34
    trend_headers = ["Month", "Total", "Telus", "Rogers"]
    for col_idx, header in enumerate(trend_headers, start=trend_col):
        ws.cell(cellular_header_row, col_idx).value = header

    for month_index in range(len(months)):
        source_row = cellular_first_row + month_index
        ws.cell(source_row, trend_col).value = f"=A{source_row}"
        ws.cell(source_row, trend_col + 2).value = (
            f"=IFERROR(B{source_row},0)+IFERROR(C{source_row},0)+IFERROR(D{source_row},0)"
        )
        ws.cell(source_row, trend_col + 3).value = f"=IFERROR(E{source_row},0)+IFERROR(F{source_row},0)"
        ws.cell(source_row, trend_col + 1).value = f"=AJ{source_row}+AK{source_row}"

    cellular_line_chart = LineChart()
    cellular_line_chart.title = ""
    set_chart_base_style(cellular_line_chart, height=14, width=30, legend_position="t")
    cellular_line_chart.add_data(
        create_reference(ws, trend_col + 1, trend_col + 3, cellular_header_row, cellular_last_row),
        titles_from_data=True,
    )
    cellular_line_chart.set_categories(
        Reference(ws, min_col=trend_col, min_row=cellular_first_row, max_row=cellular_last_row)
    )
    apply_line_colors(cellular_line_chart, [TOTAL_BLUE, TELUS_GREEN, ROGERS_RED])
    ws.add_chart(cellular_line_chart, "I35")

    # -----------------------------------------------------------------
    # Section 2: Data
    # -----------------------------------------------------------------
    data_header_row = cellular_last_row + 18
    data_first_row = data_header_row + 1

    ws.cell(data_header_row - 1, 1).value = "Gov BC Data Monthly Spend"
    style_title_cell(ws.cell(data_header_row - 1, 1), size=14)

    data_headers = ["Month", "Data (TSMA+TSMA Lite)", "Data (TELUS NGTA)", "Data (Rogers NGTA)"]
    for col_idx, header in enumerate(data_headers, start=1):
        cell = ws.cell(data_header_row, col_idx)
        cell.value = header
        style_header_cell(cell)

    for month_index, _month in enumerate(months, start=1):
        row_idx = data_first_row + month_index - 1
        match = detail_match_formula(f'"{bge_name}"', month_index)
        formulas = [
            f'={detail_index_formula("C", match)}',
            f'={detail_value_expr("E", match)}',
            f'={detail_value_expr("K", match)}',
            f'={detail_value_expr("L", match)}',
        ]

        for col_idx, formula in enumerate(formulas, start=1):
            ws.cell(row_idx, col_idx).value = formula
            if col_idx > 1:
                ws.cell(row_idx, col_idx).number_format = CURFMT

        style_data_row(ws, row_idx, 1, 4)

    data_last_row = data_first_row + march_cutoff_idx

    # Data helper is kept hidden. The old data chart remains intentionally not added.
    data_helper_col = 22
    ws.cell(data_header_row, data_helper_col).value = "Month"
    ws.cell(data_header_row, data_helper_col + 1).value = "Telus"
    ws.cell(data_header_row, data_helper_col + 2).value = "Rogers"

    for month_index, _month in enumerate(months, start=1):
        source_row = data_first_row + month_index - 1
        ws.cell(source_row, data_helper_col).value = f"=A{source_row}"
        ws.cell(source_row, data_helper_col + 1).value = f"=IFERROR(B{source_row},0)+IFERROR(C{source_row},0)"
        ws.cell(source_row, data_helper_col + 2).value = f"=IFERROR(D{source_row},0)"
        ws.cell(source_row, data_helper_col + 1).number_format = CURFMT
        ws.cell(source_row, data_helper_col + 2).number_format = CURFMT

    # -----------------------------------------------------------------
    # Section 3: Voice
    # -----------------------------------------------------------------
    voice_header_row = data_last_row + 18
    voice_first_row = voice_header_row + 1

    ws.cell(voice_header_row - 1, 1).value = "Gov BC Voice Spend by Provider"
    style_title_cell(ws.cell(voice_header_row - 1, 1), size=14)

    voice_headers = ["Month", "Voice (TSMA+TSMA Lite)", "Voice (TELUS NGTA)", "Voice (Rogers NGTA)"]
    for col_idx, header in enumerate(voice_headers, start=1):
        cell = ws.cell(voice_header_row, col_idx)
        cell.value = header
        style_header_cell(cell)

    for month_index, _month in enumerate(months, start=1):
        row_idx = voice_first_row + month_index - 1
        match = detail_match_formula(f'"{bge_name}"', month_index)
        formulas = [
            f'={detail_index_formula("C", match)}',
            f'={detail_value_expr("F", match)}',
            f'={detail_value_expr("M", match)}',
            f'={detail_value_expr("N", match)}',
        ]

        for col_idx, formula in enumerate(formulas, start=1):
            ws.cell(row_idx, col_idx).value = formula
            if col_idx > 1:
                ws.cell(row_idx, col_idx).number_format = CURFMT

        style_data_row(ws, row_idx, 1, 4)

    voice_last_row = voice_first_row + march_cutoff_idx

    voice_helper_col = 26
    ws.cell(voice_header_row, voice_helper_col).value = "Month"
    ws.cell(voice_header_row, voice_helper_col + 1).value = "Telus"
    ws.cell(voice_header_row, voice_helper_col + 2).value = "Rogers"

    for month_index, _month in enumerate(months, start=1):
        source_row = voice_first_row + month_index - 1
        ws.cell(source_row, voice_helper_col).value = f"=A{source_row}"
        ws.cell(source_row, voice_helper_col + 1).value = f"=IFERROR(B{source_row},0)+IFERROR(C{source_row},0)"
        ws.cell(source_row, voice_helper_col + 2).value = f"=IFERROR(D{source_row},0)"
        ws.cell(source_row, voice_helper_col + 1).number_format = CURFMT
        ws.cell(source_row, voice_helper_col + 2).number_format = CURFMT

    voice_chart = BarChart()
    voice_chart.type = "col"
    voice_chart.grouping = "stacked"
    voice_chart.overlap = 100
    voice_chart.gapWidth = 40
    voice_chart.title = ""
    set_chart_base_style(voice_chart, height=14, width=30, legend_position="t")
    voice_chart.add_data(
        create_reference(ws, voice_helper_col + 1, voice_helper_col + 2, voice_header_row, voice_last_row),
        titles_from_data=True,
    )
    voice_chart.set_categories(
        Reference(ws, min_col=voice_helper_col, min_row=voice_first_row, max_row=voice_last_row)
    )
    apply_bar_colors(voice_chart, [TELUS_GREEN, ROGERS_RED])
    ws.add_chart(voice_chart, f"I{voice_header_row}")

    # -----------------------------------------------------------------
    # Section 4: Summary by service tower and provider
    # -----------------------------------------------------------------
    summary_header_row = voice_last_row + 18
    summary_first_row = summary_header_row + 1
    summary_label_col = 40

    ws.cell(summary_header_row - 1, 1).value = "Gov BC Cellular, Data and Voice Monthly Spend"
    style_title_cell(ws.cell(summary_header_row - 1, 1), size=14)

    summary_headers = ["Month", "Service Tower", "Telus", "Rogers"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(summary_header_row, col_idx)
        cell.value = header
        style_header_cell(cell)

    summary_categories = ["Cellular", "Data", "Voice"]
    summary_row = summary_first_row

    for month_index, month in enumerate(months[: march_cutoff_idx + 1], start=1):
        match = detail_match_formula('"Gov BC"', month_index)

        for cat_idx, category in enumerate(summary_categories, start=1):
            row_idx = summary_row
            ws.cell(row_idx, 1).value = f'={detail_index_formula("C", match)}'
            ws.cell(row_idx, 2).value = category

            if category == "Cellular":
                telus_formula = f'={sum_detail_expr(["D", "G", "I"], match)}'
                rogers_formula = f'={sum_detail_expr(["H", "J"], match)}'
            elif category == "Data":
                telus_formula = f'={sum_detail_expr(["E", "K"], match)}'
                rogers_formula = f'={detail_value_expr("L", match)}'
            else:
                telus_formula = f'={sum_detail_expr(["F", "M"], match)}'
                rogers_formula = f'={detail_value_expr("N", match)}'

            ws.cell(row_idx, 3).value = telus_formula
            ws.cell(row_idx, 4).value = rogers_formula
            ws.cell(row_idx, summary_label_col).value = f"{month} - {category}"
            ws.cell(row_idx, summary_label_col + 1).value = month_index * 10 + cat_idx
            ws.cell(row_idx, 3).number_format = CURFMT
            ws.cell(row_idx, 4).number_format = CURFMT

            style_data_row(ws, row_idx, 1, 6)
            summary_row += 1

    summary_last_row = summary_row - 1

    summary_chart = BarChart()
    summary_chart.type = "col"
    summary_chart.grouping = "stacked"
    summary_chart.overlap = 100
    summary_chart.gapWidth = 40
    summary_chart.title = ""
    set_chart_base_style(summary_chart, height=16, width=50, legend_position="t")
    summary_chart.add_data(
        create_reference(ws, 3, 4, summary_header_row, summary_last_row),
        titles_from_data=True,
    )
    summary_chart.set_categories(
        Reference(ws, min_col=summary_label_col, min_row=summary_first_row, max_row=summary_last_row)
    )
    apply_bar_colors(summary_chart, [TELUS_GREEN, ROGERS_RED])
    ws.add_chart(summary_chart, f"I{summary_header_row}")

    hide_helper_columns(ws, cellular_helper_col, cellular_helper_col + 2)
    hide_helper_columns(ws, data_helper_col, data_helper_col + 2)
    hide_helper_columns(ws, voice_helper_col, voice_helper_col + 2)
    hide_helper_columns(ws, trend_col, trend_col + 3)
    hide_helper_columns(ws, summary_label_col, summary_label_col + 1)

    set_column_widths(
        ws,
        {
            "A": 14,
            "B": 22,
            "C": 18,
            "D": 18,
            "E": 24,
            "F": 24,
            "I": 16,
            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
            "N": 16,
            "O": 16,
        },
    )

    ws.freeze_panes = "A3"
    return ws


# =====================================================================
# README UPDATE
# =====================================================================

def update_readme(wb):
    """Append a short change note to README_BGE_Dashboard if the sheet exists."""
    if SHEET_README not in wb.sheetnames:
        return

    ws = wb[SHEET_README]
    row = ws.max_row + 1
    ws.cell(row, 1).value = "All BGE dashboard report cleanup"
    ws.cell(row, 2).value = (
        "Dashboard report script was cleaned for GitHub readiness: duplicate functions removed, "
        "sheet names standardized, chart/table formatting centralized, dropdowns repaired using "
        "BGE_List, and business rules preserved. Out of Scope and Other remain excluded. "
        "School Districts is normalized to SD. Voice, IVR, Hosted IVR, Long Distance, and "
        "Conferencing remain rolled into Voice."
    )


# =====================================================================
# MAIN
# =====================================================================

def main():
    src = find_source_file()
    out = src.with_name(OUTPUT_FILE)

    wb = load_workbook(src)

    if SHEET_SOURCE not in wb.sheetnames:
        raise ValueError(f"{SHEET_SOURCE} not found. Cannot rebuild BGE dashboard report.")

    wb_values = load_workbook(src, data_only=True)
    months, agg, fields = build_detail_data(wb_values[SHEET_SOURCE])
    last_month_idx = get_last_month_with_data(months, agg)
    bges = get_bges(wb, agg)

    sheets_to_rebuild = [
        SHEET_BGE_LIST,
        SHEET_DETAIL_DATA,
        SHEET_CATEGORY_DASHBOARD,
        SHEET_PROVIDER_DASHBOARD,
        SHEET_GOVBC_DASHBOARD,
        *OLD_SHEET_NAMES,
    ]

    for sheet_name in sheets_to_rebuild:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    rebuild_bge_list(wb, bges)
    ensure_data_sheet(wb, months, bges, agg, fields)
    rebuild_govbc_dashboard(wb, months)
    rebuild_bge_spend_by_provider_dashboard(wb, months, bges)
    rebuild_category_dashboard(wb, months, bges, last_month_idx)
    repair_dashboard_dropdowns(wb, bges)
    update_readme(wb)

    wb.active = wb.sheetnames.index(SHEET_CATEGORY_DASHBOARD)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    move_sheet_to_front(wb, SHEET_SOURCE)
    wb.save(out)

    print(f"Created: {out}")
    print(f"BGEs: {len(bges)}; months: {len(months)} ({months[0]} to {months[-1]})")
    print("Updated BGE dashboards, detail data, charts, and dropdowns.")


if __name__ == "__main__":
    main()
